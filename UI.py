# app.py
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Dict, Any, Optional, List, Tuple

import pandas as pd
import openpyxl
import streamlit as st

# Gurobi
import gurobipy as gp
from gurobipy import GRB


# =========================
# Policy (Government-set)
# =========================
@dataclass
class Policy:
    # Multipliers (how strongly each supplier dimension counts)
    env_mult: float = 1.0
    social_mult: float = 1.0
    cost_mult: float = 1.0
    strategic_mult: float = 1.0
    improvement_mult: float = 1.0
    low_quality_mult: float = 1.0

    # Penalties (hard/soft policy penalties)
    child_labor_penalty: float = 0.0
    banned_chem_penalty: float = 0.0

    def clamp_nonnegative(self) -> "Policy":
        # If you want to allow negative multipliers, remove this clamp.
        for k, v in self.__dict__.items():
            if v is None:
                setattr(self, k, 0.0)
            else:
                setattr(self, k, float(max(0.0, v)))
        return self

    def to_dict(self) -> Dict[str, float]:
        return {k: float(v) for k, v in self.__dict__.items()}


# =========================
# Min-Cost Config
# =========================
@dataclass
class MinCostConfig:
    capacity: int = 6                 # max number of matches total
    suppliers_to_select: int = 1      # choose K suppliers
    target_matches: int = 6           # require at least this many matches
    min_utility: float = 0.0          # if matched, must satisfy utility >= min_utility
    big_m: Optional[float] = None     # auto if None
    output_flag: int = 0              # Gurobi OutputFlag
    sheet_name: str = "Min Cost Agent"

    # Optional tie-break:
    # First minimize cost, then maximize total utility (policy-driven)
    lexicographic_tiebreak: bool = True


# =========================
# MinCostAgent (MILP)
# =========================
class MinCostAgent:
    """
    Minimum-Cost Matching Agent (MILP, Gurobi)

    Variables:
      y[i] ∈ {0,1}  supplier selected
      z[i,u] ∈ {0,1} user u matched to supplier i

    Constraints:
      sum_i y[i] == K
      sum_i z[i,u] <= 1                   (each user at most once)
      z[i,u] <= y[i]                       (link)
      sum_{i,u} z[i,u] <= capacity
      sum_{i,u} z[i,u] >= target_matches
      If z[i,u]=1 -> Utility(i,u) >= min_utility   (Big-M)

    Objective (primary):
      Minimize total procurement cost:
        sum_{i,u} Cost(i) * z[i,u]
      where Cost(i) is policy-weighted supplier score + penalties

    Secondary (optional):
      Maximize total utility among equal-cost solutions.
    """

    # ---------- Excel loaders ----------
    @staticmethod
    def suppliers_from_excel(
        xlsx_bytes_or_path,
        sheet_name: str = "Min Cost Agent",
        header_row: int = 3,
        first_row: int = 4,
        last_row: int = 10,
        start_col: int = 18,   # R = 18 (default based on your note: supplier table starts at R3)
    ) -> pd.DataFrame:
        wb = openpyxl.load_workbook(xlsx_bytes_or_path, data_only=True)
        ws = wb[sheet_name]

        # Expect 9 columns (Supplier + 8 attributes)
        headers = [ws.cell(header_row, c).value for c in range(start_col, start_col + 9)]
        col_map = {
            "Supplier": "supplier_id",
            "Environmental Risk": "env_risk",
            "Social Risk": "social_risk",
            "Cost Score": "cost_score",
            "Strategic Importance": "strategic",
            "Improvement Potential": "improvement",
            "Child Labor": "child_labor",
            "Banned Chemicals": "banned_chem",
            "Low Product Quality": "low_quality",
        }

        rows = []
        for r in range(first_row, last_row + 1):
            supplier_id = ws.cell(r, start_col).value
            if supplier_id is None:
                continue
            row = {}
            for i, h in enumerate(headers):
                raw = ws.cell(r, start_col + i).value
                row[col_map.get(h, h)] = raw
            rows.append(row)

        df = pd.DataFrame(rows)
        df["supplier_id"] = df["supplier_id"].astype(str)
        num_cols = ["env_risk","social_risk","cost_score","strategic","improvement","child_labor","banned_chem","low_quality"]
        for c in num_cols:
            df[c] = df[c].astype(float)
        return df

    @staticmethod
    def users_from_excel(
        xlsx_bytes_or_path,
        sheet_name: str = "Min Cost Agent",
        header_row: int = 24,
        start_col: int = 2,   # B = 2 (default based on your note: user table starts at B24)
    ) -> pd.DataFrame:
        wb = openpyxl.load_workbook(xlsx_bytes_or_path, data_only=True)
        ws = wb[sheet_name]

        # Expect 7 columns (Users + 6 weights)
        headers = [ws.cell(header_row, c).value for c in range(start_col, start_col + 7)]
        col_map = {
            "Users": "user_id",
            "Environmental Risk": "w_env",
            "Social Risk": "w_social",
            "Cost Score": "w_cost",
            "Strategic Importance": "w_strategic",
            "Improvement Potential": "w_improvement",
            "Low Product Quality": "w_low_quality",
        }

        rows = []
        r = header_row + 1
        while True:
            uid = ws.cell(r, start_col).value
            if uid is None:
                break
            row = {}
            for i, h in enumerate(headers):
                raw = ws.cell(r, start_col + i).value
                row[col_map.get(h, h)] = raw
            rows.append(row)
            r += 1

        df = pd.DataFrame(rows)
        df["user_id"] = df["user_id"].astype(str)
        for c in ["w_env","w_social","w_cost","w_strategic","w_improvement","w_low_quality"]:
            df[c] = df[c].astype(float)

        # If your Excel stores low_quality weight as negative, flip here.
        # If not needed, comment it out.
        df["w_low_quality"] = -df["w_low_quality"]
        return df

    # ---------- init ----------
    def __init__(
        self,
        suppliers_df: pd.DataFrame,
        users_df: pd.DataFrame,
        policy: Optional[Policy] = None,
        cfg: Optional[MinCostConfig] = None,
    ):
        self.suppliers = suppliers_df.copy()
        self.users = users_df.copy()
        self.policy = (policy or Policy()).clamp_nonnegative()
        self.cfg = cfg or MinCostConfig()

        self.model: Optional[gp.Model] = None
        self.y = None
        self.z = None

        self._prep()

    def _prep(self) -> None:
        s_req = {"supplier_id","env_risk","social_risk","cost_score","strategic","improvement","child_labor","banned_chem","low_quality"}
        u_req = {"user_id","w_env","w_social","w_cost","w_strategic","w_improvement","w_low_quality"}
        if not s_req.issubset(self.suppliers.columns):
            raise ValueError(f"suppliers_df eksik kolon: {s_req - set(self.suppliers.columns)}")
        if not u_req.issubset(self.users.columns):
            raise ValueError(f"users_df eksik kolon: {u_req - set(self.users.columns)}")

        self.suppliers["supplier_id"] = self.suppliers["supplier_id"].astype(str)
        self.users["user_id"] = self.users["user_id"].astype(str)

    def _auto_big_m(self) -> float:
        # Conservative M derived from magnitudes in data (utility constraint)
        p = self.policy
        s = self.suppliers
        u = self.users

        max_s = {
            "env_risk": float((p.env_mult * s["env_risk"]).abs().max()),
            "social_risk": float((p.social_mult * s["social_risk"]).abs().max()),
            "cost_score": float((p.cost_mult * s["cost_score"]).abs().max()),
            "strategic": float((p.strategic_mult * s["strategic"]).abs().max()),
            "improvement": float((p.improvement_mult * s["improvement"]).abs().max()),
            "low_quality": float((p.low_quality_mult * s["low_quality"]).abs().max()),
        }
        max_u = {
            "w_env": float(u["w_env"].abs().max()),
            "w_social": float(u["w_social"].abs().max()),
            "w_cost": float(u["w_cost"].abs().max()),
            "w_strategic": float(u["w_strategic"].abs().max()),
            "w_improvement": float(u["w_improvement"].abs().max()),
            "w_low_quality": float(u["w_low_quality"].abs().max()),
        }

        pref_max = (
            max_u["w_env"] * max_s["env_risk"] +
            max_u["w_social"] * max_s["social_risk"] +
            max_u["w_cost"] * max_s["cost_score"] +
            max_u["w_strategic"] * max_s["strategic"] +
            max_u["w_improvement"] * max_s["improvement"] +
            max_u["w_low_quality"] * max_s["low_quality"]
        )
        pen_max = float((p.child_labor_penalty * s["child_labor"] + p.banned_chem_penalty * s["banned_chem"]).max())
        return float(pref_max + pen_max + 10.0)

    def _utility(self, i: str, u: str, s_dicts: Dict[str, Dict[str, float]], u_dicts: Dict[str, Dict[str, float]]) -> float:
        p = self.policy
        s = s_dicts
        w = u_dicts

        user_pref = (
            w["w_env"][u] * (p.env_mult * s["env_risk"][i]) +
            w["w_social"][u] * (p.social_mult * s["social_risk"][i]) +
            w["w_cost"][u] * (p.cost_mult * s["cost_score"][i]) +
            w["w_strategic"][u] * (p.strategic_mult * s["strategic"][i]) +
            w["w_improvement"][u] * (p.improvement_mult * s["improvement"][i]) +
            w["w_low_quality"][u] * (p.low_quality_mult * s["low_quality"][i])
        )

        policy_pen = (
            p.child_labor_penalty * s["child_labor"][i] +
            p.banned_chem_penalty * s["banned_chem"][i]
        )

        return float(user_pref - policy_pen)

    def _cost(self, i: str, s_dicts: Dict[str, Dict[str, float]]) -> float:
        """
        Procurement cost per matched unit for supplier i.
        You can tune this mapping to match your business definition of \"cost\".
        """
        p = self.policy
        s = s_dicts

        base = (
            p.env_mult * s["env_risk"][i] +
            p.social_mult * s["social_risk"][i] +
            p.cost_mult * s["cost_score"][i] +
            p.strategic_mult * s["strategic"][i] +
            p.improvement_mult * s["improvement"][i] +
            p.low_quality_mult * s["low_quality"][i]
        )
        penalties = (
            p.child_labor_penalty * s["child_labor"][i] +
            p.banned_chem_penalty * s["banned_chem"][i]
        )
        return float(base + penalties)

    def build(self, name: str = "MinCostAgent") -> gp.Model:
        cfg = self.cfg
        p = self.policy

        Suppliers = self.suppliers["supplier_id"].tolist()
        Users = self.users["user_id"].tolist()

        # Supplier dicts
        s_dicts = {
            "env_risk": dict(zip(self.suppliers["supplier_id"], self.suppliers["env_risk"])),
            "social_risk": dict(zip(self.suppliers["supplier_id"], self.suppliers["social_risk"])),
            "cost_score": dict(zip(self.suppliers["supplier_id"], self.suppliers["cost_score"])),
            "strategic": dict(zip(self.suppliers["supplier_id"], self.suppliers["strategic"])),
            "improvement": dict(zip(self.suppliers["supplier_id"], self.suppliers["improvement"])),
            "low_quality": dict(zip(self.suppliers["supplier_id"], self.suppliers["low_quality"])),
            "child_labor": dict(zip(self.suppliers["supplier_id"], self.suppliers["child_labor"])),
            "banned_chem": dict(zip(self.suppliers["supplier_id"], self.suppliers["banned_chem"])),
        }

        # User dicts
        u_dicts = {
            "w_env": dict(zip(self.users["user_id"], self.users["w_env"])),
            "w_social": dict(zip(self.users["user_id"], self.users["w_social"])),
            "w_cost": dict(zip(self.users["user_id"], self.users["w_cost"])),
            "w_strategic": dict(zip(self.users["user_id"], self.users["w_strategic"])),
            "w_improvement": dict(zip(self.users["user_id"], self.users["w_improvement"])),
            "w_low_quality": dict(zip(self.users["user_id"], self.users["w_low_quality"])),
        }

        M = cfg.big_m if cfg.big_m is not None else self._auto_big_m()

        m = gp.Model(name)
        m.Params.OutputFlag = cfg.output_flag

        # -------------------- Variables --------------------
        y = m.addVars(Suppliers, vtype=GRB.BINARY, name="y_select_supplier")
        z = m.addVars(Suppliers, Users, vtype=GRB.BINARY, name="z_match")

        # -------------------- Constraints --------------------
        m.addConstr(gp.quicksum(y[i] for i in Suppliers) == cfg.suppliers_to_select, name="select_k_suppliers")

        for u in Users:
            m.addConstr(gp.quicksum(z[i, u] for i in Suppliers) <= 1, name=f"user_once[{u}]")

        for i in Suppliers:
            for u in Users:
                m.addConstr(z[i, u] <= y[i], name=f"link[{i},{u}]")

        m.addConstr(gp.quicksum(z[i, u] for i in Suppliers for u in Users) <= cfg.capacity, name="capacity_total")

        # Force some service level so the trivial 0-match solution is infeasible.
        m.addConstr(gp.quicksum(z[i, u] for i in Suppliers for u in Users) >= cfg.target_matches, name="target_matches")

        # Utility feasibility: if matched, utility >= min_utility
        for i in Suppliers:
            for u in Users:
                # Utility(i,u) expression
                UserPref_i_u = (
                    (u_dicts["w_env"][u] * (p.env_mult * s_dicts["env_risk"][i])) +
                    (u_dicts["w_social"][u] * (p.social_mult * s_dicts["social_risk"][i])) +
                    (u_dicts["w_cost"][u] * (p.cost_mult * s_dicts["cost_score"][i])) +
                    (u_dicts["w_strategic"][u] * (p.strategic_mult * s_dicts["strategic"][i])) +
                    (u_dicts["w_improvement"][u] * (p.improvement_mult * s_dicts["improvement"][i])) +
                    (u_dicts["w_low_quality"][u] * (p.low_quality_mult * s_dicts["low_quality"][i]))
                )
                PolicyPenalty_i = (
                    (p.child_labor_penalty * s_dicts["child_labor"][i]) +
                    (p.banned_chem_penalty * s_dicts["banned_chem"][i])
                )
                Utility_i_u = UserPref_i_u - PolicyPenalty_i

                # Utility_i_u >= min_utility - M*(1 - z[i,u])
                m.addConstr(Utility_i_u >= cfg.min_utility - M * (1 - z[i, u]), name=f"utility[{i},{u}]")

        # -------------------- Objective --------------------
        # Cost(i) per matched unit (supplier-level)
        cost_i = {i: self._cost(i, s_dicts) for i in Suppliers}

        TotalCost = gp.quicksum(cost_i[i] * z[i, u] for i in Suppliers for u in Users)

        if not cfg.lexicographic_tiebreak:
            m.setObjective(TotalCost, GRB.MINIMIZE)
        else:
            # Min cost first, then maximize total utility
            TotalUtility = gp.quicksum(
                (  # same utility expression but multiplied by z
                    (
                        (u_dicts["w_env"][u] * (p.env_mult * s_dicts["env_risk"][i])) +
                        (u_dicts["w_social"][u] * (p.social_mult * s_dicts["social_risk"][i])) +
                        (u_dicts["w_cost"][u] * (p.cost_mult * s_dicts["cost_score"][i])) +
                        (u_dicts["w_strategic"][u] * (p.strategic_mult * s_dicts["strategic"][i])) +
                        (u_dicts["w_improvement"][u] * (p.improvement_mult * s_dicts["improvement"][i])) +
                        (u_dicts["w_low_quality"][u] * (p.low_quality_mult * s_dicts["low_quality"][i]))
                    )
                    - (
                        (p.child_labor_penalty * s_dicts["child_labor"][i]) +
                        (p.banned_chem_penalty * s_dicts["banned_chem"][i])
                    )
                ) * z[i, u]
                for i in Suppliers
                for u in Users
            )

            m.setObjectiveN(TotalCost, index=0, priority=2, name="min_total_cost")
            m.setObjectiveN(TotalUtility, index=1, priority=1, name="max_total_utility")

        self.model, self.y, self.z = m, y, z
        return m

    def solve(self) -> Dict[str, Any]:
        if self.model is None:
            self.build()

        self.model.optimize()
        if self.model.Status not in (GRB.OPTIMAL, GRB.SUBOPTIMAL):
            raise RuntimeError(f"No solution. Status={self.model.Status}")

        Suppliers = self.suppliers["supplier_id"].tolist()
        Users = self.users["user_id"].tolist()
        # Rebuild dicts for reporting
        s_dicts = {
            "env_risk": dict(zip(self.suppliers["supplier_id"], self.suppliers["env_risk"])),
            "social_risk": dict(zip(self.suppliers["supplier_id"], self.suppliers["social_risk"])),
            "cost_score": dict(zip(self.suppliers["supplier_id"], self.suppliers["cost_score"])),
            "strategic": dict(zip(self.suppliers["supplier_id"], self.suppliers["strategic"])),
            "improvement": dict(zip(self.suppliers["supplier_id"], self.suppliers["improvement"])),
            "low_quality": dict(zip(self.suppliers["supplier_id"], self.suppliers["low_quality"])),
            "child_labor": dict(zip(self.suppliers["supplier_id"], self.suppliers["child_labor"])),
            "banned_chem": dict(zip(self.suppliers["supplier_id"], self.suppliers["banned_chem"])),
        }
        u_dicts = {
            "w_env": dict(zip(self.users["user_id"], self.users["w_env"])),
            "w_social": dict(zip(self.users["user_id"], self.users["w_social"])),
            "w_cost": dict(zip(self.users["user_id"], self.users["w_cost"])),
            "w_strategic": dict(zip(self.users["user_id"], self.users["w_strategic"])),
            "w_improvement": dict(zip(self.users["user_id"], self.users["w_improvement"])),
            "w_low_quality": dict(zip(self.users["user_id"], self.users["w_low_quality"])),
        }

        chosen = [i for i in Suppliers if self.y[i].X > 0.5]

        pairs: List[Tuple[str, str]] = [(u, i) for i in Suppliers for u in Users if self.z[i, u].X > 0.5]
        match_df = pd.DataFrame(pairs, columns=["user_id", "supplier_id"])

        if len(match_df) > 0:
            match_df["utility"] = match_df.apply(lambda r: self._utility(r["supplier_id"], r["user_id"], s_dicts, u_dicts), axis=1)
            match_df["unit_cost"] = match_df["supplier_id"].map(lambda i: self._cost(i, s_dicts))
            match_df = match_df.sort_values(["supplier_id", "user_id"]).reset_index(drop=True)
        else:
            match_df["utility"] = []
            match_df["unit_cost"] = []

        total_cost = float(sum(match_df["unit_cost"])) if len(match_df) else 0.0
        avg_utility = float(match_df["utility"].mean()) if len(match_df) else 0.0

        return {
            "status": int(self.model.Status),
            "objective_value": float(self.model.ObjVal),
            "total_cost": total_cost,
            "avg_utility": avg_utility,
            "chosen_suppliers": chosen,
            "num_matched": int(len(pairs)),
            "matches": match_df,
            "policy": self.policy.to_dict(),
            "cfg": self.cfg,
        }


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="Min-Cost Policy Matching", layout="wide")

st.title("Minimum Cost Matching • Government Policy Selector")

with st.expander("Ne yapıyor?", expanded=False):
    st.markdown(
        """
- Excel’den supplier + user tablolarını okur  
- Hükümet politikası parametrelerini dropdown ile seçtirir  
- Gurobi ile **minimum-cost** matching çözer  
- Seçilen supplier(lar), match’ler ve **utility** değerlerini gösterir  
        """
    )

# ---------- helpers ----------
def _parse_candidates(text: str) -> List[float]:
    vals = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        vals.append(float(part))
    # unique + sorted
    vals = sorted(set(vals))
    return vals

@st.cache_data(show_spinner=False)
def load_tables_from_excel(
    file_bytes: bytes,
    sheet_name: str,
    s_header_row: int,
    s_first_row: int,
    s_last_row: int,
    s_start_col: int,
    u_header_row: int,
    u_start_col: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    bio = io.BytesIO(file_bytes)
    suppliers = MinCostAgent.suppliers_from_excel(
        bio, sheet_name=sheet_name,
        header_row=s_header_row, first_row=s_first_row, last_row=s_last_row, start_col=s_start_col
    )
    bio2 = io.BytesIO(file_bytes)
    users = MinCostAgent.users_from_excel(
        bio2, sheet_name=sheet_name,
        header_row=u_header_row, start_col=u_start_col
    )
    return suppliers, users

def policy_selector(prefix: str = "pol") -> Policy:
    st.subheader("Government Policy (dropdown ile seç)")

    default_candidates = "0, 0.25, 0.5, 1, 2, 5, 10"
    cand_text = st.text_input(
        "Dropdown seçenekleri (virgülle ayır)",
        value=default_candidates,
        help="Tüm policy parametre dropdown’ları bu listeden değer seçer."
    )
    candidates = _parse_candidates(cand_text) or [0.0, 1.0]

    cols = st.columns(3)
    with cols[0]:
        env = st.selectbox("env_mult", candidates, index=min(3, len(candidates)-1), key=f"{prefix}_env")
        social = st.selectbox("social_mult", candidates, index=min(3, len(candidates)-1), key=f"{prefix}_social")
        cost = st.selectbox("cost_mult", candidates, index=min(3, len(candidates)-1), key=f"{prefix}_cost")
    with cols[1]:
        strategic = st.selectbox("strategic_mult", candidates, index=min(3, len(candidates)-1), key=f"{prefix}_strategic")
        improvement = st.selectbox("improvement_mult", candidates, index=min(3, len(candidates)-1), key=f"{prefix}_improvement")
        lq = st.selectbox("low_quality_mult", candidates, index=min(3, len(candidates)-1), key=f"{prefix}_lq")
    with cols[2]:
        child = st.selectbox("child_labor_penalty", candidates, index=0, key=f"{prefix}_child")
        banned = st.selectbox("banned_chem_penalty", candidates, index=0, key=f"{prefix}_banned")

    pol = Policy(
        env_mult=env,
        social_mult=social,
        cost_mult=cost,
        strategic_mult=strategic,
        improvement_mult=improvement,
        low_quality_mult=lq,
        child_labor_penalty=child,
        banned_chem_penalty=banned,
    ).clamp_nonnegative()

    st.caption("Seçili policy (JSON):")
    st.json(pol.to_dict())
    return pol


# ---------- sidebar controls ----------
st.sidebar.header("Excel & Model Ayarları")

uploaded = st.sidebar.file_uploader("Excel yükle (.xlsx)", type=["xlsx"])

sheet_name = st.sidebar.text_input("Sheet name", value="Min Cost Agent")

st.sidebar.subheader("Supplier table koordinatları")
s_header_row = st.sidebar.number_input("Supplier header row", min_value=1, value=3, step=1)
s_first_row = st.sidebar.number_input("Supplier first data row", min_value=1, value=4, step=1)
s_last_row = st.sidebar.number_input("Supplier last data row", min_value=1, value=10, step=1)
s_start_col = st.sidebar.number_input("Supplier start col (A=1, B=2, ... R=18)", min_value=1, value=18, step=1)

st.sidebar.subheader("User table koordinatları")
u_header_row = st.sidebar.number_input("User header row", min_value=1, value=24, step=1)
u_start_col = st.sidebar.number_input("User start col (A=1, B=2, ...)", min_value=1, value=2, step=1)

st.sidebar.subheader("Optimization config")
capacity = st.sidebar.number_input("capacity (max matches)", min_value=0, value=6, step=1)
suppliers_to_select = st.sidebar.number_input("K (suppliers_to_select)", min_value=1, value=1, step=1)

min_utility = st.sidebar.number_input("min_utility (match olunca)", value=0.0, step=0.5)
target_matches = st.sidebar.number_input("target_matches (en az)", min_value=0, value=6, step=1)

lexi = st.sidebar.checkbox("Lexicographic tie-break (min cost → max utility)", value=True)

output_flag = st.sidebar.checkbox("Gurobi log göster (OutputFlag=1)", value=False)

# ---------- main ----------
colA, colB = st.columns([1.15, 0.85], gap="large")

with colB:
    policy = policy_selector()

with colA:
    st.subheader("Veri & Çözüm")

    if uploaded is None:
        st.info("Devam etmek için Excel dosyasını yükle.")
        st.stop()

    try:
        suppliers_df, users_df = load_tables_from_excel(
            uploaded.getvalue(),
            sheet_name=sheet_name,
            s_header_row=int(s_header_row),
            s_first_row=int(s_first_row),
            s_last_row=int(s_last_row),
            s_start_col=int(s_start_col),
            u_header_row=int(u_header_row),
            u_start_col=int(u_start_col),
        )
    except Exception as e:
        st.error("Excel okuma hatası. Sheet adı ve koordinatları kontrol et.")
        st.exception(e)
        st.stop()

    with st.expander("Suppliers (okunan tablo)", expanded=False):
        st.dataframe(suppliers_df, use_container_width=True)

    with st.expander("Users (okunan tablo)", expanded=False):
        st.dataframe(users_df, use_container_width=True)

    cfg = MinCostConfig(
        capacity=int(capacity),
        suppliers_to_select=int(suppliers_to_select),
        target_matches=int(target_matches),
        min_utility=float(min_utility),
        output_flag=1 if output_flag else 0,
        sheet_name=sheet_name,
        lexicographic_tiebreak=bool(lexi),
    )

    solve_btn = st.button("Solve (Gurobi)", type="primary")

    if solve_btn:
        try:
            agent = MinCostAgent(suppliers_df, users_df, policy=policy, cfg=cfg)
            agent.build()
            sol = agent.solve()
        except gp.GurobiError as ge:
            st.error("Gurobi hatası (lisans / environment olabilir).")
            st.exception(ge)
            st.stop()
        except Exception as e:
            st.error("Model solve hatası.")
            st.exception(e)
            st.stop()

        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("Matched", sol["num_matched"])
        kpi2.metric("Total cost", f"{sol['total_cost']:.3f}")
        kpi3.metric("Avg utility", f"{sol['avg_utility']:.3f}")
        kpi4.metric("Obj (Gurobi)", f"{sol['objective_value']:.3f}")

        st.markdown("### Seçilen supplier(lar)")
        st.write(sol["chosen_suppliers"])

        st.markdown("### Match tablosu (utility & unit_cost)")
        st.dataframe(sol["matches"], use_container_width=True)

        st.markdown("### Hızlı özet")
        if sol["num_matched"] > 0:
            by_supplier = sol["matches"].groupby("supplier_id").agg(
                matched=("user_id", "count"),
                avg_utility=("utility", "mean"),
                avg_unit_cost=("unit_cost", "mean"),
                total_cost=("unit_cost", "sum"),
            ).reset_index()
            st.dataframe(by_supplier, use_container_width=True)
        else:
            st.warning("Hiç match çıkmadı. target_matches / min_utility / capacity / policy değerlerini kontrol et.")
