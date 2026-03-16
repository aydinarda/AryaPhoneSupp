from __future__ import annotations

import argparse
import random
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

SERVER_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = Path(__file__).resolve().parents[3]

if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

from app.matching_engine import run_market_matching
from app.service import evaluate_manual, get_tables
from app.settings import FIXED_POLICY, GAME_SETTINGS


DEFAULT_OUTPUT = PROJECT_ROOT / "random_session_matching_report.xlsx"
DEFAULT_ROUND_COUNT = 10
DEFAULT_TEAM_COUNT = 30
DEFAULT_SEED = 20260312


def _parse_bool_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return float(value) != 0.0
    text = str(value).strip().lower()
    if text in {"true", "t", "1", "yes", "y"}:
        return True
    if text in {"false", "f", "0", "no", "n", "", "none", "null"}:
        return False
    return False


def _safe_sheet_title(title: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in r'[]:*?/\\')
    return cleaned[:31] or "Sheet"


def _write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    if df is None or df.empty:
        ws.cell(row=start_row, column=start_col, value="No data")
        return start_row + 1

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row=start_row + row_offset, column=col_idx, value=value)

    return start_row + len(df) + 2


def _autosize_columns(ws, max_width: int = 24) -> None:
    for column_cells in ws.columns:
        length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(length + 2, 10), max_width)


def _build_submission_rows(run_no: int, team_count: int, supplier_ids: list[str], rng: random.Random) -> pd.DataFrame:
    start_time = datetime(2026, 3, 12, 9, 0, 0) + timedelta(minutes=run_no * 10)
    teams = [f"Team_{idx:02d}" for idx in range(1, team_count + 1)]
    rng.shuffle(teams)

    rows: list[dict[str, Any]] = []
    max_pick_count = min(8, len(supplier_ids))

    for order_idx, team in enumerate(teams):
        pick_count = rng.randint(1, max_pick_count)
        picks = rng.sample(supplier_ids, pick_count)
        manual = evaluate_manual("max_profit", picks)
        metrics = manual["metrics"]
        created_at = start_time + timedelta(seconds=order_idx)

        rows.append(
            {
                "run_no": run_no,
                "submission_order": order_idx + 1,
                "team": team,
                "player_name": team,
                "num_suppliers": pick_count,
                "selected_suppliers": ",".join(picks),
                "feasible": bool(manual.get("feasible", False)),
                "profit": float(metrics.get("profit_total", 0.0)),
                "utility": float(metrics.get("utility_total", 0.0)),
                "env_avg": float(metrics.get("avg_env", 0.0)),
                "social_avg": float(metrics.get("avg_social", 0.0)),
                "cost_avg": float(metrics.get("avg_cost", 0.0)),
                "strategic_avg": float(metrics.get("avg_strategic", 0.0)),
                "improvement_avg": float(metrics.get("avg_improvement", 0.0)),
                "low_quality_avg": float(metrics.get("avg_low_quality", 0.0)),
                "created_at": created_at.isoformat(),
            }
        )

    return pd.DataFrame(rows).sort_values("submission_order").reset_index(drop=True)


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def _build_team_profiles(submissions_df: pd.DataFrame, suppliers_df: pd.DataFrame) -> tuple[dict[str, dict[str, Any]], list[str]]:
    suppliers = suppliers_df.copy()
    suppliers["supplier_id"] = suppliers["supplier_id"].astype(str)
    supplier_by_id = {
        str(row["supplier_id"]): {
            "env_risk": row.get("env_risk"),
            "social_risk": row.get("social_risk"),
            "cost_score": row.get("cost_score"),
            "strategic": row.get("strategic"),
            "improvement": row.get("improvement"),
            "low_quality": row.get("low_quality"),
            "child_labor": row.get("child_labor"),
            "banned_chem": row.get("banned_chem"),
        }
        for _, row in suppliers.iterrows()
    }

    profiles: dict[str, dict[str, Any]] = {}
    excluded: list[str] = []

    for row in submissions_df.itertuples(index=False):
        team = str(row.team)
        picks = [item.strip() for item in str(row.selected_suppliers).split(",") if item.strip()]
        picks = list(dict.fromkeys(picks))
        valid = [pid for pid in picks if pid in supplier_by_id]
        if not valid:
            excluded.append(team)
            continue

        selected = [supplier_by_id[pid] for pid in valid]
        count = float(len(selected))
        avg_env = sum(_safe_float(r.get("env_risk")) for r in selected) / count
        avg_social = sum(_safe_float(r.get("social_risk")) for r in selected) / count
        avg_cost = sum(_safe_float(r.get("cost_score")) for r in selected) / count
        avg_strategic = sum(_safe_float(r.get("strategic")) for r in selected) / count
        avg_improvement = sum(_safe_float(r.get("improvement")) for r in selected) / count
        avg_low_quality = sum(_safe_float(r.get("low_quality")) for r in selected) / count
        avg_child_labor = sum(_safe_float(r.get("child_labor")) for r in selected) / count
        avg_banned_chem = sum(_safe_float(r.get("banned_chem")) for r in selected) / count

        if not (
            avg_env <= float(GAME_SETTINGS.env_cap) + 1e-12
            and avg_social <= float(GAME_SETTINGS.social_cap) + 1e-12
        ):
            excluded.append(team)
            continue

        profiles[team] = {
            "team": team,
            "created_at": str(row.created_at),
            "picked_suppliers": valid,
            "avg_env": avg_env,
            "avg_social": avg_social,
            "avg_cost": avg_cost,
            "avg_strategic": avg_strategic,
            "avg_improvement": avg_improvement,
            "avg_low_quality": avg_low_quality,
            "avg_child_labor": avg_child_labor,
            "avg_banned_chem": avg_banned_chem,
        }

    return profiles, sorted(set(excluded))


def _run_matching_for_submissions(submissions_df: pd.DataFrame, suppliers_df: pd.DataFrame, users_df: pd.DataFrame, market_capacity: int, user_limit: int | None = None) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    feasible_mask = submissions_df["feasible"].map(_parse_bool_flag)
    eligible_df = submissions_df[feasible_mask].copy().reset_index(drop=True)
    excluded_teams = submissions_df.loc[~feasible_mask, "team"].tolist()

    if eligible_df.empty:
        return {
            "market_to_users": {},
            "user_to_market": {},
            "unmatched_users": [],
            "market_loads": {},
            "excluded_infeasible_users": excluded_teams,
            "meta": {
                "solver": "stable_gale_shapley",
                "user_count": 0,
                "eligible_user_count": 0,
                "submitted_user_count": len(submissions_df),
                "matched_count": 0,
                "unmatched_count": 0,
                "market_option_count": 0,
                "total_capacity": 0,
                "infeasible_excluded_count": len(excluded_teams),
            },
        }, {}

    team_profiles, profile_excluded = _build_team_profiles(eligible_df, suppliers_df)
    excluded_teams = sorted(set(excluded_teams + profile_excluded))

    if not team_profiles:
        return {
            "market_to_users": {},
            "user_to_market": {},
            "unmatched_users": [],
            "market_loads": {},
            "excluded_infeasible_users": excluded_teams,
            "excluded_infeasible_teams": excluded_teams,
            "matching_target": "team_product",
            "meta": {
                "solver": "stable_gale_shapley",
                "user_count": 0,
                "eligible_team_count": 0,
                "submitted_team_count": len(submissions_df),
                "matched_count": 0,
                "unmatched_count": 0,
                "market_option_count": 0,
                "total_capacity": 0,
                "infeasible_excluded_count": len(excluded_teams),
                "user_pool_count": 0,
            },
        }, {}

    served_df = users_df.copy().reset_index(drop=True)
    if user_limit is not None and user_limit < len(served_df):
        served_df = served_df.head(user_limit).reset_index(drop=True)
    team_ids = sorted(team_profiles.keys())

    user_score_map: dict[str, dict[str, float]] = {}
    users_payload: list[dict[str, Any]] = []

    for _, u in served_df.iterrows():
        user_id = str(u.get("user_id", "")).strip()
        if not user_id:
            continue

        w_env = _safe_float(u.get("w_env"))
        w_social = _safe_float(u.get("w_social"))
        w_str = _safe_float(u.get("w_strategic"))
        w_imp = _safe_float(u.get("w_improvement"))
        w_lq = _safe_float(u.get("w_low_quality"))

        utilities: dict[str, float] = {}
        for team_id in team_ids:
            profile = team_profiles[team_id]
            score = (
                w_env * (float(FIXED_POLICY.env_mult) * (5.0 - float(profile["avg_env"])))
                + w_social * (float(FIXED_POLICY.social_mult) * (5.0 - float(profile["avg_social"])))
                + w_str * (float(FIXED_POLICY.strategic_mult) * (float(profile["avg_strategic"]) - 1.0))
                + w_imp * (float(FIXED_POLICY.improvement_mult) * (float(profile["avg_improvement"]) - 1.0))
                + w_lq * (float(FIXED_POLICY.low_quality_mult) * (5.0 - float(profile["avg_low_quality"])))
            )
            utilities[team_id] = float(score)

        choices = sorted(team_ids, key=lambda tid: (-utilities[tid], tid))
        users_payload.append({"user_id": user_id, "choices": choices, "utilities": utilities})
        user_score_map[user_id] = utilities

    market_options = []
    for team_id in team_ids:
        priority = sorted(
            [str(u.get("user_id", "")) for u in users_payload],
            key=lambda uid: (-user_score_map.get(uid, {}).get(team_id, 0.0), uid),
        )
        market_options.append(
            {
                "option_id": team_id,
                "capacity": int(market_capacity),
                "priority": priority,
                "request_time": team_profiles[team_id].get("created_at"),
            }
        )

    result = run_market_matching(users=users_payload, market_options=market_options)
    result["excluded_infeasible_users"] = excluded_teams
    result["excluded_infeasible_teams"] = excluded_teams
    result["matching_target"] = "team_product"
    result["meta"]["eligible_team_count"] = len(team_profiles)
    result["meta"]["submitted_team_count"] = len(submissions_df)
    result["meta"]["user_pool_count"] = len(users_payload)
    result["meta"]["infeasible_excluded_count"] = len(excluded_teams)
    return result, team_profiles


def _augment_with_matching(submissions_df: pd.DataFrame, matching: dict[str, Any]) -> pd.DataFrame:
    df = submissions_df.copy()
    market_to_users = matching.get("market_to_users", {}) or {}
    excluded = set(matching.get("excluded_infeasible_users", []) or [])
    df["matching_eligible"] = ~df["team"].isin(excluded) & df["feasible"].astype(bool)
    df["matched_user_count"] = df["team"].map(lambda team: len(market_to_users.get(str(team), []) or []))
    df["matched_user_ids"] = df["team"].map(
        lambda team: ", ".join(sorted(market_to_users.get(str(team), []) or []))
    )
    df["matched"] = (df["matched_user_count"] > 0) & df["matching_eligible"]
    df["excluded_reason"] = df.apply(
        lambda row: "infeasible" if not bool(row.feasible) else "",
        axis=1,
    )
    return df


def _build_round_financial_rows(
    run_no: int,
    submissions_df: pd.DataFrame,
    matching: dict[str, Any],
    team_profiles: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    market_to_users = matching.get("market_to_users", {}) or {}
    rows: list[dict[str, Any]] = []

    for row in submissions_df.itertuples(index=False):
        team = str(row.team)
        profile = team_profiles.get(team)
        matched_users = [str(uid) for uid in (market_to_users.get(team) or [])]
        matched_count = len(matched_users)

        if profile is None:
            unit_margin = 0.0
            cost_component = 0.0
            penalty = 0.0
        else:
            avg_cost = float(profile.get("avg_cost", 0.0))
            avg_child_labor = float(profile.get("avg_child_labor", 0.0))
            avg_banned_chem = float(profile.get("avg_banned_chem", 0.0))
            cost_component = float(GAME_SETTINGS.cost_scale) * avg_cost
            penalty = (
                float(FIXED_POLICY.child_labor_penalty) * avg_child_labor
                + float(FIXED_POLICY.banned_chem_penalty) * avg_banned_chem
            )
            unit_margin = float(GAME_SETTINGS.price_per_user) - cost_component - penalty

        realized_profit = float(matched_count) * unit_margin
        rows.append(
            {
                "run_no": run_no,
                "team": team,
                "matched_user_count": matched_count,
                "matched_users": ", ".join(matched_users),
                "sale_price_per_user": float(GAME_SETTINGS.price_per_user),
                "cost_component_per_user": float(cost_component),
                "penalty_per_user": float(penalty),
                "unit_margin": float(unit_margin),
                "realized_profit_round": float(realized_profit),
            }
        )

    return pd.DataFrame(rows)


def _build_matching_assignment_rows(
    run_no: int,
    matching: dict[str, Any],
    team_profiles: dict[str, dict[str, Any]] | None = None,
) -> pd.DataFrame:
    user_to_market = matching.get("user_to_market", {}) or {}
    unmatched = set(matching.get("unmatched_users", []) or [])
    rows = []
    for user_id in sorted(user_to_market):
        destination = user_to_market.get(user_id)
        suppliers_str = ""
        if destination and team_profiles and destination in team_profiles:
            suppliers_str = ", ".join(team_profiles[destination].get("picked_suppliers", []))
        rows.append(
            {
                "run_no": run_no,
                "user_id": user_id,
                "matched_team_product": destination,
                "team_selected_suppliers": suppliers_str,
                "matched": user_id not in unmatched and destination is not None,
            }
        )
    return pd.DataFrame(rows)


def _build_market_rows(run_no: int, matching: dict[str, Any]) -> pd.DataFrame:
    market_to_users = matching.get("market_to_users", {}) or {}
    market_loads = matching.get("market_loads", {}) or {}
    rows = []
    option_ids = sorted(set(market_to_users.keys()) | set(market_loads.keys()))
    for option_id in option_ids:
        users = market_to_users.get(option_id, []) or []
        load = market_loads.get(option_id, {}) or {}
        rows.append(
            {
                "run_no": run_no,
                "team_product": option_id,
                "assigned_count": len(users or []),
                "capacity": load.get("capacity"),
                "remaining_capacity": load.get("remaining_capacity"),
                "assigned_users": ", ".join(users or []),
            }
        )
    return pd.DataFrame(rows)


def _build_run_summary(run_no: int, market_capacity: int, submissions_df: pd.DataFrame, matching: dict[str, Any]) -> dict[str, Any]:
    meta = matching.get("meta", {}) or {}
    matched_count = int(meta.get("matched_count", 0) or 0)
    total_teams = len(submissions_df)
    eligible_team_count = int(meta.get("eligible_team_count", 0) or 0)
    user_pool_count = int(meta.get("user_pool_count", meta.get("user_count", 0)) or 0)
    return {
        "run_no": run_no,
        "market_capacity": market_capacity,
        "solver": meta.get("solver", "unknown"),
        "team_count": total_teams,
        "eligible_team_count": eligible_team_count,
        "user_pool_count": user_pool_count,
        "infeasible_excluded_count": int(meta.get("infeasible_excluded_count", 0) or 0),
        "matched_user_count": matched_count,
        "unmatched_user_count": int(meta.get("unmatched_count", max(0, user_pool_count - matched_count)) or 0),
        "user_match_rate": round(matched_count / user_pool_count, 4) if user_pool_count else 0.0,
        "feasible_submission_count": int(submissions_df["feasible"].sum()),
        "total_profit": float(submissions_df["profit"].sum()),
        "total_utility": float(submissions_df["utility"].sum()),
        "avg_profit": float(submissions_df["profit"].mean()),
        "avg_utility": float(submissions_df["utility"].mean()),
        "supplier_options_used": int(submissions_df["selected_suppliers"].str.split(",").explode().nunique()),
    }


def _create_leaderboard_plot(run_df: pd.DataFrame, out_path: Path, title: str) -> None:
    fig, ax = plt.subplots(figsize=(9, 5.2))
    color_map = {True: "#059669", False: "#dc2626"}

    for row in run_df.itertuples(index=False):
        is_eligible = bool(getattr(row, "matching_eligible", False))
        alpha = 0.85 if is_eligible else 0.35
        ax.scatter(
            row.profit,
            row.utility,
            color=color_map[bool(row.matched)],
            marker="o",
            s=70,
            alpha=alpha,
            edgecolors="white",
            linewidths=0.5,
        )
        ax.annotate(str(row.team), (row.profit, row.utility), fontsize=7, alpha=0.75, xytext=(4, 3), textcoords="offset points")

    ax.set_title(title)
    ax.set_xlabel("Profit")
    ax.set_ylabel("Utility")
    ax.grid(True, alpha=0.2)

    legend_labels = [
        "Green = matched",
        "Red = unmatched",
        "Faded = infeasible, excluded",
    ]
    ax.text(1.01, 0.5, "\n".join(legend_labels), transform=ax.transAxes, va="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(out_path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _create_round_progress_plot(
    progress_df: pd.DataFrame,
    out_path: Path,
    title: str,
    top_n_labeled: int = 10,
) -> None:
    """Plot each team's cumulative profit progression across rounds.

    y-axis: round number (1 → N), x-axis: cumulative realized profit (total gained).
    Top-N teams by final profit are highlighted and labeled; rest are shown faded.
    """
    teams = sorted(progress_df["team"].unique())
    final_profits = (
        progress_df.groupby("team")["cumulative_profit"]
        .max()
        .sort_values(ascending=False)
    )
    top_teams = set(final_profits.head(top_n_labeled).index.tolist())

    cmap = plt.get_cmap("tab20")
    color_map = {team: cmap((i % 20) / 20.0) for i, team in enumerate(sorted(teams))}

    fig, ax = plt.subplots(figsize=(12, 7))

    for team in teams:
        team_data = progress_df[progress_df["team"] == team].sort_values("run_no")
        x = team_data["cumulative_profit"].values
        y = team_data["run_no"].values
        is_top = team in top_teams
        color = color_map[team] if is_top else "#bbbbbb"
        alpha = 0.90 if is_top else 0.25
        lw = 2.2 if is_top else 0.8
        zorder = 3 if is_top else 1
        ax.plot(x, y, color=color, alpha=alpha, linewidth=lw, marker="o", markersize=3.5, zorder=zorder)
        if is_top:
            ax.annotate(
                str(team),
                (x[-1], y[-1]),
                fontsize=7.5,
                color=color,
                xytext=(6, 0),
                textcoords="offset points",
                va="center",
                fontweight="bold",
            )

    rounds = sorted(progress_df["run_no"].unique())
    ax.set_yticks(rounds)
    ax.set_yticklabels([f"Round {r:02d}" for r in rounds])
    ax.set_xlabel("Total Gained — Cumulative Realized Profit", fontsize=11)
    ax.set_ylabel("Round", fontsize=11)
    ax.set_title(title, fontsize=13, fontweight="bold")
    ax.grid(True, axis="x", alpha=0.25, linestyle="--")
    ax.grid(True, axis="y", alpha=0.15)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    fig.tight_layout()
    fig.savefig(out_path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _populate_run_sheet(ws, run_summary: dict[str, Any], run_df: pd.DataFrame, market_df: pd.DataFrame, image_path: Path) -> None:
    ws["A1"] = f"Round {int(run_summary['run_no']):02d} Summary"
    ws["A1"].font = Font(bold=True, size=14)

    summary_df = pd.DataFrame([run_summary])
    next_row = _write_dataframe(ws, summary_df, start_row=3)

    ws.cell(row=next_row, column=1, value="Leaderboard Rows").font = Font(bold=True)
    next_row = _write_dataframe(ws, run_df, start_row=next_row + 1)

    ws.cell(row=3, column=16, value="Team Product Assignments").font = Font(bold=True)
    _write_dataframe(ws, market_df, start_row=4, start_col=16)

    img = XLImage(str(image_path))
    img.width = 900
    img.height = 500
    ws.add_image(img, "P12")
    ws.freeze_panes = "A4"
    _autosize_columns(ws)


def generate_report(output_path: Path, rounds: int, team_count: int, seed: int, user_limit: int | None = None) -> Path:
    suppliers_df, users_df = get_tables()
    suppliers_df = suppliers_df.copy()
    suppliers_df["supplier_id"] = suppliers_df["supplier_id"].astype(str)
    supplier_ids = suppliers_df["supplier_id"].tolist()
    rng = random.Random(seed)

    run_summaries: list[dict[str, Any]] = []
    all_submissions: list[pd.DataFrame] = []
    all_assignments: list[pd.DataFrame] = []
    all_market_rows: list[pd.DataFrame] = []
    all_round_financials: list[pd.DataFrame] = []
    run_frames: dict[int, pd.DataFrame] = {}
    run_market_frames: dict[int, pd.DataFrame] = {}
    image_paths: dict[int, Path] = {}
    cumulative_profit_by_team = {f"Team_{idx:02d}": 0.0 for idx in range(1, team_count + 1)}
    cumulative_matched_by_team = {f"Team_{idx:02d}": 0 for idx in range(1, team_count + 1)}

    with tempfile.TemporaryDirectory(prefix="arya_random_report_") as temp_dir_str:
        temp_dir = Path(temp_dir_str)

        for run_no in range(1, rounds + 1):
            market_capacity = int(GAME_SETTINGS.default_market_capacity)
            submissions_df = _build_submission_rows(run_no, team_count, supplier_ids, rng)
            matching, team_profiles = _run_matching_for_submissions(submissions_df, suppliers_df, users_df, market_capacity, user_limit=user_limit)
            augmented_df = _augment_with_matching(submissions_df, matching)
            round_financial_df = _build_round_financial_rows(run_no, augmented_df, matching, team_profiles)

            round_profit_map = {
                str(r.team): float(r.realized_profit_round)
                for r in round_financial_df.itertuples(index=False)
            }
            round_matched_map = {
                str(r.team): int(r.matched_user_count)
                for r in round_financial_df.itertuples(index=False)
            }

            augmented_df["realized_profit_round"] = augmented_df["team"].map(lambda t: round_profit_map.get(str(t), 0.0))
            augmented_df["matched_user_count"] = augmented_df["team"].map(lambda t: round_matched_map.get(str(t), 0))
            augmented_df["cumulative_profit"] = augmented_df["team"].map(lambda t: cumulative_profit_by_team.get(str(t), 0.0)) + augmented_df["realized_profit_round"]
            augmented_df["cumulative_matched_users"] = augmented_df["team"].map(lambda t: cumulative_matched_by_team.get(str(t), 0)) + augmented_df["matched_user_count"]

            for row in augmented_df.itertuples(index=False):
                team = str(row.team)
                cumulative_profit_by_team[team] = float(getattr(row, "cumulative_profit", cumulative_profit_by_team.get(team, 0.0)))
                cumulative_matched_by_team[team] = int(getattr(row, "cumulative_matched_users", cumulative_matched_by_team.get(team, 0)))

            assignment_df = _build_matching_assignment_rows(run_no, matching, team_profiles)
            market_df = _build_market_rows(run_no, matching)
            summary = _build_run_summary(run_no, market_capacity, augmented_df, matching)
            summary["round_realized_profit_total"] = float(round_financial_df["realized_profit_round"].sum()) if len(round_financial_df) else 0.0

            image_path = temp_dir / f"run_{run_no:02d}_leaderboard.png"
            _create_leaderboard_plot(augmented_df, image_path, f"Round {run_no:02d} Leaderboard")

            run_summaries.append(summary)
            all_submissions.append(augmented_df)
            all_assignments.append(assignment_df)
            all_market_rows.append(market_df)
            all_round_financials.append(round_financial_df)
            run_frames[run_no] = augmented_df
            run_market_frames[run_no] = market_df
            image_paths[run_no] = image_path

        summary_df = pd.DataFrame(run_summaries)
        submissions_df = pd.concat(all_submissions, ignore_index=True)
        assignments_df = pd.concat(all_assignments, ignore_index=True)
        market_rows_df = pd.concat(all_market_rows, ignore_index=True)
        round_financials_df = pd.concat(all_round_financials, ignore_index=True)

        team_match_counts = (
            submissions_df.groupby("team", as_index=False)
            .agg(
                eligible_runs=("matching_eligible", "sum"),
                matched_runs=("matched", "sum"),
                infeasible_runs=("feasible", lambda s: int((~s.astype(bool)).sum())),
                total_runs=("team", "size"),
            )
            .assign(match_rate_when_eligible=lambda df: df.apply(lambda row: row["matched_runs"] / row["eligible_runs"] if row["eligible_runs"] else 0.0, axis=1))
            .sort_values(["matched_runs", "eligible_runs", "team"], ascending=[False, False, True])
        )

        team_metrics = (
            submissions_df.groupby("team", as_index=False)
            .agg(
                avg_profit=("profit", "mean"),
                avg_utility=("utility", "mean"),
                feasible_runs=("feasible", "sum"),
                avg_supplier_count=("num_suppliers", "mean"),
            )
        )
        team_overview_df = team_match_counts.merge(team_metrics, on="team", how="left")

        final_game_results_df = (
            submissions_df.groupby("team", as_index=False)
            .agg(
                total_realized_profit=("realized_profit_round", "sum"),
                total_matched_users=("matched_user_count", "sum"),
                feasible_rounds=("feasible", "sum"),
                avg_submission_profit=("profit", "mean"),
                avg_submission_utility=("utility", "mean"),
            )
            .sort_values(["total_realized_profit", "total_matched_users", "team"], ascending=[False, False, True])
            .reset_index(drop=True)
        )
        final_game_results_df.insert(0, "final_rank", range(1, len(final_game_results_df) + 1))

        supplier_overview_df = (
            market_rows_df.groupby("team_product", as_index=False)
            .agg(
                total_assigned_users=("assigned_count", "sum"),
                avg_assigned_users=("assigned_count", "mean"),
                max_assigned_users=("assigned_count", "max"),
                total_remaining_capacity=("remaining_capacity", "sum"),
                avg_capacity=("capacity", "mean"),
                run_count=("run_no", "size"),
            )
            .sort_values(["total_assigned_users", "team_product"], ascending=[False, True])
        )

        # Round-by-round progress chart (y=round, x=cumulative profit)
        progress_image_path = temp_dir / "round_progress.png"
        _create_round_progress_plot(
            submissions_df[["run_no", "team", "cumulative_profit"]].copy(),
            progress_image_path,
            f"Round-by-Round Progression — {rounds} Rounds, {team_count} Teams",
        )

        workbook = Workbook()
        overview_ws = workbook.active
        overview_ws.title = "Overview"
        overview_ws["A1"] = "Random Multi-round Matching Overview"
        overview_ws["A1"].font = Font(bold=True, size=14)
        overview_ws["A2"] = f"Seed: {seed}"
        overview_ws["B2"] = f"Rounds: {rounds}"
        overview_ws["C2"] = f"Teams per round: {team_count}"
        effective_users = user_limit if user_limit is not None else len(users_df)
        overview_ws["D2"] = f"User pool: {effective_users}"
        _write_dataframe(overview_ws, summary_df, start_row=4)
        overview_ws.freeze_panes = "A5"
        _autosize_columns(overview_ws)

        team_ws = workbook.create_sheet("Team Match Counts")
        team_ws["A1"] = "Per-team match performance across all rounds"
        team_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(team_ws, team_overview_df, start_row=3)
        team_ws.freeze_panes = "A4"
        _autosize_columns(team_ws)

        final_ws = workbook.create_sheet("Final Game Results")
        final_ws["A1"] = "Final standings after all rounds"
        final_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(final_ws, final_game_results_df, start_row=3)
        final_ws.freeze_panes = "A4"
        _autosize_columns(final_ws)

        supplier_ws = workbook.create_sheet("Team Product Match Counts")
        supplier_ws["A1"] = "Per-team-product assigned user counts across all rounds"
        supplier_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(supplier_ws, supplier_overview_df, start_row=3)
        supplier_ws.freeze_panes = "A4"
        _autosize_columns(supplier_ws)

        submissions_ws = workbook.create_sheet("All Submissions")
        submissions_ws["A1"] = "All simulated submissions"
        submissions_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(submissions_ws, submissions_df, start_row=3)
        submissions_ws.freeze_panes = "A4"
        _autosize_columns(submissions_ws)

        assignments_ws = workbook.create_sheet("Assignments")
        assignments_ws["A1"] = "User-to-team-product matching assignments"
        assignments_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(assignments_ws, assignments_df, start_row=3)
        assignments_ws.freeze_panes = "A4"
        _autosize_columns(assignments_ws)

        market_ws = workbook.create_sheet("Team Product Loads")
        market_ws["A1"] = "Team-product assignment loads"
        market_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(market_ws, market_rows_df, start_row=3)
        market_ws.freeze_panes = "A4"
        _autosize_columns(market_ws)

        financial_ws = workbook.create_sheet("Round Financials")
        financial_ws["A1"] = "Per-round realized financial results"
        financial_ws["A1"].font = Font(bold=True, size=14)
        _write_dataframe(financial_ws, round_financials_df, start_row=3)
        financial_ws.freeze_panes = "A4"
        _autosize_columns(financial_ws)

        progress_ws = workbook.create_sheet("Round Progress")
        progress_ws["A1"] = "Cumulative profit progression per team across rounds (y=round, x=total gained)"
        progress_ws["A1"].font = Font(bold=True, size=14)
        prog_img = XLImage(str(progress_image_path))
        prog_img.width = 1100
        prog_img.height = 700
        progress_ws.add_image(prog_img, "A3")
        _autosize_columns(progress_ws)

        for run_no in range(1, rounds + 1):
            ws = workbook.create_sheet(_safe_sheet_title(f"Round {run_no:02d}"))
            _populate_run_sheet(ws, run_summaries[run_no - 1], run_frames[run_no], run_market_frames[run_no], image_paths[run_no])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a random, multi-round matching Excel report for one game.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output XLSX path")
    parser.add_argument("--rounds", type=int, default=DEFAULT_ROUND_COUNT, help="Number of rounds in the single simulation")
    parser.add_argument("--runs", type=int, default=None, help="Deprecated alias of --rounds")
    parser.add_argument("--teams", type=int, default=DEFAULT_TEAM_COUNT, help="Team count per round")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED, help="Random seed for reproducible rounds")
    parser.add_argument("--users", type=int, default=None, help="Limit user pool size (default: all users)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rounds = int(args.runs) if args.runs is not None else int(args.rounds)
    output = generate_report(output_path=args.output, rounds=rounds, team_count=args.teams, seed=args.seed, user_limit=args.users)
    print(output)


if __name__ == "__main__":
    main()