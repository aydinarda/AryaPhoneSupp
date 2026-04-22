"""
classroom_sim.py
----------------
30-student classroom simulation, 10 rounds.

Decision logic (applied at the start of each round based on the previous round):
  1. Below mean profit + previously infeasible  -> find a feasible combination
  2. Below mean profit + audit-excluded         -> 50% chance: swap suppliers, 50%: keep
  3. Below mean profit + feasible               -> cut price 5-15% (stay profitable)
  4. Market share > 8.5%                        -> 20% chance raise price, 10% chance lower

Audit: probability=0.30, catch=0.80

Output: classroom_simulation.xlsx
  - "Round 1" ... "Round 10"  -> data table
                                  + pie chart: market share (existing)
                                  + scatter: Buyer Utility vs Market Utility
                                  + scatter: Profit vs Buyer Utility
                                  + scatter: Profit vs Market Utility
  - "Cumulative"              -> cumulative profit table
                                  + cumulative profit line chart (existing)
                                  + round-by-round profit line chart (existing)
                                  + market share evolution line chart (existing)
                                  + cumulative buyer utility line chart
                                  + cumulative market utility line chart
                                  + scatter: cumulative Buyer Utility vs Market Utility
                                  + scatter: cumulative Profit vs Buyer Utility
                                  + scatter: cumulative Profit vs Market Utility
"""

from __future__ import annotations

import sys
import random
import warnings
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

# ── Path setup ────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))
warnings.filterwarnings("ignore")

from server.app.audit import run_audit
from server.app.beta_density import BetaDensity
from server.app.customer_segment import CustomerSegment
from server.app.mnl_market import BuyerProfile, run_mnl_market
from server.app.service import get_tables
from server.app.settings import GAME_SETTINGS

# ── Constants ─────────────────────────────────────────────────────────────────
N_STUDENTS           = 30
N_ROUNDS             = 10
BASE_PRICE           = float(GAME_SETTINGS.price_per_user)
COST_SCALE           = float(GAME_SETTINGS.cost_scale)
ENV_CAP              = float(GAME_SETTINGS.env_cap)
SOC_CAP              = float(GAME_SETTINGS.social_cap)
DELTA                = float(GAME_SETTINGS.price_sensitivity_delta)
AUDIT_PROB           = 0.30
CATCH_PROB           = 0.80
HIGH_SHARE_THRESHOLD = 0.085

RNG         = random.Random(42)
OUTPUT_PATH = Path(__file__).parent / "classroom_simulation.xlsx"

# ── Load supplier / user data ─────────────────────────────────────────────────
suppliers_df, users_df = get_tables()
N_USERS      = len(users_df)
supplier_ids = suppliers_df["supplier_id"].tolist()

HAS_CATEGORIES = (
    "category" in suppliers_df.columns
    and suppliers_df["category"].notna().any()
)

cat_groups: dict[str, list[str]] = {}
if HAS_CATEGORIES:
    for _, row in suppliers_df.iterrows():
        cat = str(row["category"]).strip()
        if cat and cat.lower() != "nan":
            cat_groups.setdefault(cat, []).append(str(row["supplier_id"]))

print(f"Suppliers: {len(supplier_ids)}  |  Users: {N_USERS}  |  Categorical: {HAS_CATEGORIES}")
if HAS_CATEGORIES:
    for cat, ids in cat_groups.items():
        print(f"  {cat}: {ids}")

# ── Supplier helpers ───────────────────────────────────────────────────────────

def avg_attrs(picks: list[str]) -> dict[str, float]:
    sel = suppliers_df[suppliers_df["supplier_id"].isin(picks)]
    if sel.empty:
        return {}
    cols = ["env_risk", "social_risk", "cost_score", "strategic"]
    for c in ("child_labor", "banned_chem"):
        if c in suppliers_df.columns:
            cols.append(c)
    return {c: float(sel[c].mean()) for c in cols if c in sel.columns}


def is_feasible(picks: list[str]) -> bool:
    if not picks:
        return False
    a = avg_attrs(picks)
    return (
        a.get("env_risk", 99)    <= ENV_CAP + 1e-9
        and a.get("social_risk", 99) <= SOC_CAP + 1e-9
        and _category_ok(picks)
    )


def _category_ok(picks: list[str]) -> bool:
    if not HAS_CATEGORIES:
        return True
    pick_set = set(picks)
    return all(
        sum(1 for sid in ids if sid in pick_set) == 1
        for ids in cat_groups.values()
    )


def min_profitable_price(picks: list[str]) -> float:
    """Lowest price that keeps unit margin positive (cost + $0.50 buffer)."""
    a = avg_attrs(picks)
    return COST_SCALE * a.get("cost_score", 0.0) + 0.50


def random_picks(rng: random.Random) -> list[str]:
    if HAS_CATEGORIES:
        return [rng.choice(ids) for ids in cat_groups.values()]
    k = rng.randint(1, min(4, len(supplier_ids)))
    return rng.sample(supplier_ids, k)


def random_feasible_picks(rng: random.Random, max_attempts: int = 300) -> list[str]:
    """Random feasible selection. Falls back to last attempt if none found."""
    last = random_picks(rng)
    for _ in range(max_attempts):
        picks = random_picks(rng)
        if is_feasible(picks):
            return picks
        last = picks
    return last


# ── Student agent ─────────────────────────────────────────────────────────────

@dataclass
class Student:
    name:               str
    picks:              list[str]
    price:              float
    prev_profit:        float = 0.0
    prev_demand_share:  float = 0.0
    prev_feasible:      bool  = True
    prev_audit_excl:    bool  = False
    played:             bool  = False

    def decide(self, mean_profit: float, rng: random.Random) -> None:
        """Update picks / price before this round, based on the previous round."""
        if not self.played:
            return

        below_mean = self.prev_profit < mean_profit
        high_share = self.prev_demand_share > HIGH_SHARE_THRESHOLD

        # Priority 1: below mean + was infeasible -> find feasible combo
        if below_mean and not self.prev_feasible:
            self.picks = random_feasible_picks(rng)
            self.price = max(min_profitable_price(self.picks) + 5.0, BASE_PRICE)
            return

        # Priority 2: below mean + audit excluded -> 50/50 swap
        if below_mean and self.prev_audit_excl:
            if rng.random() < 0.50:
                self.picks = random_feasible_picks(rng)
                self.price = max(min_profitable_price(self.picks) + 5.0, self.price)
            return

        # Priority 3: below mean + feasible -> cut price (stay profitable)
        if below_mean and self.prev_feasible:
            cut = rng.uniform(0.05, 0.15)
            new_price = self.price * (1.0 - cut)
            self.price = max(new_price, min_profitable_price(self.picks))
            return

        # Priority 4: high market share -> price adjustment
        if high_share:
            r = rng.random()
            if r < 0.20:
                self.price *= rng.uniform(1.05, 1.15)
            elif r < 0.30:
                new_price = self.price * rng.uniform(0.90, 0.95)
                self.price = max(new_price, min_profitable_price(self.picks))


# ── Initialise students ────────────────────────────────────────────────────────
students: list[Student] = []
for i in range(1, N_STUDENTS + 1):
    picks = random_picks(RNG)
    price = BASE_PRICE + RNG.uniform(-10.0, 10.0)
    students.append(Student(name=f"Team_{i:02d}", picks=picks, price=round(price, 2)))


# ── Market segment builder ────────────────────────────────────────────────────

def make_segments() -> list[CustomerSegment]:
    bd = BetaDensity(alpha=3, beta=3)
    us = users_df.sort_values("w_cost").reset_index(drop=True)
    return [
        CustomerSegment(
            segment_id=str(row["user_id"]),
            density=float(bd.density_at((i + 0.5) / N_USERS)),
            w_env=float(row.get("w_env", 0.0)),
            w_social=float(row.get("w_social", 0.0)),
            w_cost=float(row.get("w_cost", 1.0)),
            w_low_quality=float(row.get("w_low_quality", 0.0)),
        )
        for i, (_, row) in enumerate(us.iterrows())
    ]


# ── Run simulation ─────────────────────────────────────────────────────────────
all_rounds: list[pd.DataFrame] = []

for round_no in range(1, N_ROUNDS + 1):
    print(f"\n--- Round {round_no} -------------------------------------------")

    if round_no > 1:
        prev_df     = all_rounds[-1]
        mean_profit = prev_df["realized_profit"].mean()
        for student in students:
            student.decide(mean_profit, RNG)

    # Build feasible team profiles
    team_profiles: dict[str, dict[str, Any]] = {}
    for s in students:
        if is_feasible(s.picks):
            a = avg_attrs(s.picks)
            team_profiles[s.name] = {
                "team":             s.name,
                "price_per_user":   s.price,
                "picked_suppliers": s.picks[:],
                "avg_env":          a.get("env_risk",    0.0),
                "avg_social":       a.get("social_risk", 0.0),
                "avg_cost":         a.get("cost_score",  0.0),
                "avg_strategic":    a.get("strategic",   0.0),
            }

    infeasible_names = {s.name for s in students if s.name not in team_profiles}

    # Audit
    audit_result   = run_audit(
        team_profiles=team_profiles,
        suppliers_df=suppliers_df,
        audit_probability=AUDIT_PROB,
        catch_probability=CATCH_PROB,
        rng=RNG,
    )
    audit_penalty_map = {team: float(penalty) for team, penalty in audit_result.team_penalties.items()}
    audit_pen_set = set(audit_result.penalized_teams)

    print(
        f"  Active: {len(team_profiles):2d}  |  "
        f"Infeasible: {len(infeasible_names):2d}  |  "
        f"Audit penalized: {len(audit_pen_set):2d}"
    )

    # MNL market
    active_profiles = [
        BuyerProfile(
            team_name=tp["team"],
            price_per_user=tp["price_per_user"],
            avg_env=tp["avg_env"],
            avg_social=tp["avg_social"],
            utility_adjustment=audit_penalty_map.get(tp["team"], 0.0),
        )
        for tp in team_profiles.values()
    ]
    segments   = make_segments()
    mnl_result = run_mnl_market(active_profiles, segments, delta=DELTA, u_outside=None) \
        if active_profiles else None

    # Collect round results
    rows = []
    for student in students:
        a             = avg_attrs(student.picks)
        feasible      = is_feasible(student.picks)
        audit_excl    = student.name in audit_excl_set
        avg_env       = a.get("env_risk",    0.0)
        avg_social    = a.get("social_risk", 0.0)
        avg_strategic = a.get("strategic",   0.0)

        if not feasible or audit_excl or mnl_result is None:
            demand_share    = 0.0
            realized_profit = 0.0
            market_utility  = 0.0
        else:
            br              = mnl_result.buyer_results.get(student.name)
            demand_share    = br.total_demand if br else 0.0
            cost_per_unit   = COST_SCALE * a.get("cost_score", 0.0)
            unit_margin     = student.price - cost_per_unit
            realized_profit = demand_share * N_USERS * unit_margin
            market_utility  = round((br.realized_utility * N_USERS) if br else 0.0, 3)

        buyer_utility = round(
            0.1 * realized_profit
            + (5.0 - avg_env)
            + (5.0 - avg_social)
            + (5.0 - avg_strategic),
            4,
        )

        rows.append({
            "round_no":         round_no,
            "team":             student.name,
            "picks":            ",".join(student.picks),
            "price":            round(student.price, 2),
            "feasible":         feasible,
            "audit_excluded":   audit_excl,
            "avg_env":          round(avg_env,       3),
            "avg_social":       round(avg_social,    3),
            "avg_cost":         round(a.get("cost_score", 0.0), 3),
            "avg_strategic":    round(avg_strategic, 3),
            "demand_share_pct": round(demand_share * 100, 2),
            "realized_profit":  round(realized_profit, 2),
            "market_utility":   market_utility,
            "buyer_utility":    buyer_utility,
        })

        student.prev_profit       = realized_profit
        student.prev_demand_share = demand_share
        student.prev_feasible     = feasible
        student.prev_audit_excl   = audit_excl
        student.played            = True

    round_df = pd.DataFrame(rows)
    all_rounds.append(round_df)

    top3 = round_df.nlargest(3, "demand_share_pct")
    for _, r in top3.iterrows():
        print(
            f"  {r['team']}: {r['demand_share_pct']:.1f}%  "
            f"profit=${r['realized_profit']:.0f}  "
            f"buyer_u={r['buyer_utility']:.2f}"
        )

full_df = pd.concat(all_rounds, ignore_index=True)


# ── Colour palette ────────────────────────────────────────────────────────────
PALETTE = [
    "#2563eb", "#dc2626", "#059669", "#7c3aed", "#d97706", "#0891b2",
    "#be123c", "#65a30d", "#9333ea", "#ea580c", "#0284c7", "#16a34a",
    "#b91c1c", "#7e22ce", "#ca8a04", "#0369a1", "#15803d", "#c2410c",
    "#6d28d9", "#b45309", "#075985", "#166534", "#991b1b", "#5b21b6",
    "#92400e", "#0c4a6e", "#14532d", "#7f1d1d", "#4c1d95", "#78350f",
]
teams  = [s.name for s in students]
colors = {team: PALETTE[i % len(PALETTE)] for i, team in enumerate(teams)}


# ── Excel helpers ──────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1e3a5f")
HDR_FONT  = Font(bold=True, color="FFFFFF")
RED_FILL  = PatternFill("solid", fgColor="FEE2E2")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
CTR       = Alignment(horizontal="center")

def _header(ws, row: int, cols: list[str]) -> None:
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, CTR

def _fig_to_image(fig: plt.Figure, dpi: int = 120) -> XLImage:
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return XLImage(buf)

def _scatter_ax(
    ax: plt.Axes,
    xvals: list[float],
    yvals: list[float],
    labels: list[str],
    clrs: list[str],
    xlabel: str,
    ylabel: str,
    title: str,
) -> None:
    for x, y, lbl, c in zip(xvals, yvals, labels, clrs):
        ax.scatter(x, y, color=c, s=55, zorder=3, edgecolors="white", linewidths=0.4)
        short = lbl.replace("Team_", "T")
        ax.annotate(short, (x, y), fontsize=5.5, ha="center", va="bottom",
                    xytext=(0, 3), textcoords="offset points")
    ax.set_xlabel(xlabel, fontsize=9)
    ax.set_ylabel(ylabel, fontsize=9)
    ax.set_title(title, fontsize=9, fontweight="bold")
    ax.grid(alpha=0.2)


# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

ROUND_COLS = [
    "Team", "Picks", "Price", "Feasible", "Audit Excl.",
    "Avg Env", "Avg Social", "Avg Cost", "Strategic",
    "Share %", "Profit", "Mkt Utility", "Buyer Utility",
]
N_COLS = len(ROUND_COLS)  # 13

for round_no, rdf in enumerate(all_rounds, 1):
    ws = wb.create_sheet(title=f"Round {round_no}")
    _header(ws, 1, ROUND_COLS)

    for ri, (_, row) in enumerate(rdf.iterrows(), 2):
        ws.cell(ri,  1, row["team"])
        ws.cell(ri,  2, row["picks"])
        ws.cell(ri,  3, row["price"])
        ws.cell(ri,  4, "Yes" if row["feasible"] else "No")
        ws.cell(ri,  5, "Yes" if row["audit_excluded"] else "")
        ws.cell(ri,  6, row["avg_env"])
        ws.cell(ri,  7, row["avg_social"])
        ws.cell(ri,  8, row["avg_cost"])
        ws.cell(ri,  9, row["avg_strategic"])
        ws.cell(ri, 10, row["demand_share_pct"])
        ws.cell(ri, 11, row["realized_profit"])
        ws.cell(ri, 12, row["market_utility"])
        ws.cell(ri, 13, row["buyer_utility"])

        fill = (RED_FILL  if not row["feasible"] else
                WARN_FILL if row["audit_excluded"] else None)
        if fill:
            for ci in range(1, N_COLS + 1):
                ws.cell(ri, ci).fill = fill

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for col_letter in ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
        ws.column_dimensions[col_letter].width = 13

    # Only teams that actually participated in the market
    active = rdf[rdf["demand_share_pct"] > 0].copy()

    # Chart grid: 2x2 below the data table (table rows 1..N_STUDENTS+1)
    ROW1 = N_STUDENTS + 4   # 34 — top row of charts
    ROW2 = ROW1 + 33        # 67 — bottom row of charts

    # ── Pie chart: market share (original) ───────────────────────────────────
    if not active.empty:
        large = active[active["demand_share_pct"] >= 2.0].copy()
        small = active[active["demand_share_pct"] <  2.0]
        labels = large["team"].tolist()
        sizes  = large["demand_share_pct"].tolist()
        clrs   = [colors[t] for t in labels]
        if not small.empty:
            labels.append(f"Others ({len(small)})")
            sizes.append(float(small["demand_share_pct"].sum()))
            clrs.append("#94a3b8")

        fig, ax = plt.subplots(figsize=(7, 5))
        wedges, _, autotexts = ax.pie(
            sizes, labels=None,
            autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
            colors=clrs, startangle=140, pctdistance=0.78,
        )
        for at in autotexts:
            at.set_fontsize(7)
        ax.legend(wedges, labels, title="Team",
                  loc="center left", bbox_to_anchor=(1.0, 0.5),
                  fontsize=7, title_fontsize=8)
        ax.set_title(
            f"Round {round_no} — Market Share Distribution\n"
            f"(Active teams: {len(active)}  |  Audit p={AUDIT_PROB})",
            fontsize=10, fontweight="bold",
        )
        plt.tight_layout()
        img_pie = _fig_to_image(fig)
        img_pie.anchor = f"A{ROW1}"
        ws.add_image(img_pie)

    # ── Scatter: Buyer Utility vs Market Utility ──────────────────────────────
    if not active.empty:
        fig_s1, ax_s1 = plt.subplots(figsize=(6.5, 5))
        _scatter_ax(
            ax_s1,
            active["market_utility"].tolist(),
            active["buyer_utility"].tolist(),
            active["team"].tolist(),
            [colors[t] for t in active["team"]],
            "Market Utility",
            "Buyer Utility",
            f"Round {round_no} — Buyer Utility vs Market Utility",
        )
        plt.tight_layout()
        img_s1 = _fig_to_image(fig_s1)
        img_s1.anchor = f"M{ROW1}"
        ws.add_image(img_s1)

    # ── Scatter: Profit vs Buyer Utility ─────────────────────────────────────
    if not active.empty:
        fig_s2, ax_s2 = plt.subplots(figsize=(6.5, 5))
        _scatter_ax(
            ax_s2,
            active["realized_profit"].tolist(),
            active["buyer_utility"].tolist(),
            active["team"].tolist(),
            [colors[t] for t in active["team"]],
            "Profit ($)",
            "Buyer Utility",
            f"Round {round_no} — Profit vs Buyer Utility",
        )
        plt.tight_layout()
        img_s2 = _fig_to_image(fig_s2)
        img_s2.anchor = f"A{ROW2}"
        ws.add_image(img_s2)

    # ── Scatter: Profit vs Market Utility ────────────────────────────────────
    if not active.empty:
        fig_s3, ax_s3 = plt.subplots(figsize=(6.5, 5))
        _scatter_ax(
            ax_s3,
            active["realized_profit"].tolist(),
            active["market_utility"].tolist(),
            active["team"].tolist(),
            [colors[t] for t in active["team"]],
            "Profit ($)",
            "Market Utility",
            f"Round {round_no} — Profit vs Market Utility",
        )
        plt.tight_layout()
        img_s3 = _fig_to_image(fig_s3)
        img_s3.anchor = f"M{ROW2}"
        ws.add_image(img_s3)


# ── Cumulative sheet ───────────────────────────────────────────────────────────
ws_cum = wb.create_sheet(title="Cumulative")

pivot_profit = full_df.pivot_table(index="round_no", columns="team", values="realized_profit",  fill_value=0.0)
pivot_share  = full_df.pivot_table(index="round_no", columns="team", values="demand_share_pct", fill_value=0.0)
pivot_buyer  = full_df.pivot_table(index="round_no", columns="team", values="buyer_utility",    fill_value=0.0)
pivot_market = full_df.pivot_table(index="round_no", columns="team", values="market_utility",   fill_value=0.0)

cum_profit = pivot_profit.cumsum()
cum_buyer  = pivot_buyer.cumsum()
cum_market = pivot_market.cumsum()

# ── Cumulative profit table (teams as rows, rounds as columns) ────────────────
cum_by_team = cum_profit.T.copy()
cum_by_team["Total"] = cum_by_team.iloc[:, -1]
cum_by_team = cum_by_team.sort_values("Total", ascending=False)

table_header = ["Team"] + [f"After R{r}" for r in range(1, N_ROUNDS + 1)] + ["Cumulative Total"]
_header(ws_cum, 1, table_header)

for ri, (team, row_data) in enumerate(cum_by_team.iterrows(), 2):
    ws_cum.cell(ri, 1, team)
    for ci, r in enumerate(range(1, N_ROUNDS + 1), 2):
        ws_cum.cell(ri, ci, round(row_data.get(r, 0.0), 2))
    ws_cum.cell(ri, N_ROUNDS + 2, round(float(row_data["Total"]), 2))

ws_cum.column_dimensions["A"].width = 11
for ci in range(2, N_ROUNDS + 3):
    ws_cum.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 13

# ── Shared helpers for cumulative charts ──────────────────────────────────────
final_rank  = cum_profit.iloc[-1].sort_values(ascending=False)
top3_teams  = final_rank.head(3).index.tolist()
bot3_teams  = final_rank.tail(3).index.tolist()
highlighted = set(top3_teams + bot3_teams)
rounds_axis = list(range(1, N_ROUNDS + 1))

def _line_chart(
    pivot: pd.DataFrame,
    ylabel: str,
    title: str,
    figsize: tuple = (13, 6),
) -> plt.Figure:
    fig, ax = plt.subplots(figsize=figsize)
    for team in pivot.columns:
        vals = pivot[team].tolist()
        if team in highlighted:
            ax.plot(rounds_axis, vals, color=colors[team], linewidth=2.2,
                    alpha=1.0, label=team, zorder=3, marker="o", markersize=4)
        else:
            ax.plot(rounds_axis, vals, color="#94a3b8", linewidth=0.8,
                    alpha=0.35, zorder=1)
    ax.set_xlabel("Round", fontsize=11)
    ax.set_ylabel(ylabel, fontsize=11)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.set_xticks(rounds_axis)
    ax.legend(fontsize=9, loc="upper left", framealpha=0.9)
    ax.grid(axis="y", alpha=0.25)
    plt.tight_layout()
    return fig

def _cum_scatter(
    cum_x: pd.DataFrame,
    cum_y: pd.DataFrame,
    xlabel: str,
    ylabel: str,
    title: str,
) -> plt.Figure:
    """Scatter of each team's cumulative totals at the end of the simulation."""
    fig, ax = plt.subplots(figsize=(9, 6))
    for team in cum_x.columns:
        x = float(cum_x[team].iloc[-1])
        y = float(cum_y[team].iloc[-1])
        c = colors.get(team, "#94a3b8")
        is_top = team in top3_teams
        is_bot = team in bot3_teams
        marker = "*" if is_top else ("v" if is_bot else "o")
        size   = 120 if (is_top or is_bot) else 55
        ax.scatter(x, y, color=c, s=size, marker=marker, zorder=3,
                   edgecolors="white", linewidths=0.5)
        short = team.replace("Team_", "T")
        fs    = 7   if (is_top or is_bot) else 5.5
        fw    = "bold" if (is_top or is_bot) else "normal"
        ax.annotate(short, (x, y), fontsize=fs, fontweight=fw,
                    ha="center", va="bottom", xytext=(0, 4),
                    textcoords="offset points")
    ax.set_xlabel(xlabel, fontsize=11)
    ax.set_ylabel(ylabel, fontsize=11)
    ax.set_title(title, fontsize=11, fontweight="bold")
    ax.grid(alpha=0.2)
    plt.tight_layout()
    return fig

# Chart start row on Cumulative sheet
chart_row = N_STUDENTS + 4  # 34

# ── Chart 1: Cumulative profit line (original) ────────────────────────────────
fig1 = _line_chart(
    cum_profit,
    "Cumulative Profit ($)",
    f"Cumulative Profit — {N_STUDENTS} Students, {N_ROUNDS} Rounds\n"
    "(Top 3 & Bottom 3 by final ranking highlighted)",
)
img1 = _fig_to_image(fig1, dpi=130)
img1.anchor = f"A{chart_row}"
ws_cum.add_image(img1)

# ── Chart 2: Round-by-round profit line (original) ────────────────────────────
fig2 = _line_chart(
    pivot_profit,
    "Profit ($)",
    f"Round-by-Round Profit — {N_STUDENTS} Students, {N_ROUNDS} Rounds\n"
    "(Top 3 & Bottom 3 in final cumulative ranking highlighted)",
)
img2 = _fig_to_image(fig2, dpi=130)
img2.anchor = f"A{chart_row + 33}"
ws_cum.add_image(img2)

# ── Chart 3: Market share evolution (original) ────────────────────────────────
avg_share  = pivot_share.mean().sort_values(ascending=False)
top6_share = avg_share.head(6).index.tolist()

fig3, ax3 = plt.subplots(figsize=(13, 5))
for team in top6_share:
    vals = pivot_share[team].tolist()
    ax3.plot(rounds_axis, vals, color=colors[team], linewidth=2.0, alpha=1.0,
             label=f"{team} (avg {avg_share[team]:.1f}%)", marker="s", markersize=4)
ax3.axhline(HIGH_SHARE_THRESHOLD * 100, color="#dc2626", linestyle="--",
            linewidth=1.2, alpha=0.7,
            label=f"High-share threshold ({HIGH_SHARE_THRESHOLD * 100:.1f}%)")
ax3.set_xlabel("Round", fontsize=11)
ax3.set_ylabel("Market Share (%)", fontsize=11)
ax3.set_title("Market Share Evolution — Top 6 Teams by Average Share",
              fontsize=12, fontweight="bold")
ax3.set_xticks(rounds_axis)
ax3.legend(fontsize=9, loc="upper right", framealpha=0.9)
ax3.grid(axis="y", alpha=0.25)
plt.tight_layout()

img3 = _fig_to_image(fig3, dpi=130)
img3.anchor = f"A{chart_row + 66}"
ws_cum.add_image(img3)

# ── Chart 4: Cumulative buyer utility line ────────────────────────────────────
fig4 = _line_chart(
    cum_buyer,
    "Cumulative Buyer Utility",
    f"Cumulative Buyer Utility — {N_STUDENTS} Students, {N_ROUNDS} Rounds\n"
    "(Top 3 & Bottom 3 by profit highlighted)",
)
img4 = _fig_to_image(fig4, dpi=130)
img4.anchor = f"A{chart_row + 99}"
ws_cum.add_image(img4)

# ── Chart 5: Cumulative market utility line ───────────────────────────────────
fig5 = _line_chart(
    cum_market,
    "Cumulative Market Utility",
    f"Cumulative Market Utility — {N_STUDENTS} Students, {N_ROUNDS} Rounds\n"
    "(Top 3 & Bottom 3 by profit highlighted)",
)
img5 = _fig_to_image(fig5, dpi=130)
img5.anchor = f"A{chart_row + 132}"
ws_cum.add_image(img5)

# ── Chart 6: Cumulative scatter — Buyer Utility vs Market Utility ─────────────
fig6 = _cum_scatter(
    cum_market,
    cum_buyer,
    "Cumulative Market Utility",
    "Cumulative Buyer Utility",
    "Cumulative: Buyer Utility vs Market Utility  (* top 3  v bottom 3 by profit)",
)
img6 = _fig_to_image(fig6, dpi=130)
img6.anchor = f"A{chart_row + 165}"
ws_cum.add_image(img6)

# ── Chart 7: Cumulative scatter — Profit vs Buyer Utility ─────────────────────
fig7 = _cum_scatter(
    cum_profit,
    cum_buyer,
    "Cumulative Profit ($)",
    "Cumulative Buyer Utility",
    "Cumulative: Profit vs Buyer Utility  (* top 3  v bottom 3 by profit)",
)
img7 = _fig_to_image(fig7, dpi=130)
img7.anchor = f"A{chart_row + 198}"
ws_cum.add_image(img7)

# ── Chart 8: Cumulative scatter — Profit vs Market Utility ────────────────────
fig8 = _cum_scatter(
    cum_profit,
    cum_market,
    "Cumulative Profit ($)",
    "Cumulative Market Utility",
    "Cumulative: Profit vs Market Utility  (* top 3  v bottom 3 by profit)",
)
img8 = _fig_to_image(fig8, dpi=130)
img8.anchor = f"A{chart_row + 231}"
ws_cum.add_image(img8)

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")
print(f"\nFinal cumulative profit ranking:")
for rank, (team, profit) in enumerate(final_rank.items(), 1):
    print(f"  {rank:2d}. {team}: ${profit:,.0f}")
