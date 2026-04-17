"""
classroom_sim_compare.py
------------------------
Comparative run: same 30-student, 10-round simulation executed twice —
  Run A: BetaDensity(alpha=3.0, beta=3.0)
  Run B: BetaDensity(alpha=1.5, beta=3.2)

Both runs use the same RNG seed (42) so student initialisation and
audit draws start identically; they diverge naturally as demand shares
(and therefore profits) differ due to the different density weighting.

Each round sheet contains three stacked tables:
  1. Run A results
  2. Run B results
  3. Difference  (B minus A, numeric columns only)

Followed by side-by-side charts (pie A | pie B, then scatter comparisons).

Cumulative sheet contains line / scatter charts overlaying both runs.

Output: classroom_simulation_compare.xlsx  (same folder as this script)
"""

from __future__ import annotations

import sys
import random
import warnings
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

# ── Path setup ────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))
warnings.filterwarnings("ignore")

from server.app.audit import run_audit
from server.app.beta_density import BetaDensity
from server.app.customer_segment import CustomerSegment
from server.app.mnl_market import BuyerProfile, run_mnl_market
from server.app.service import get_tables
from server.app.settings import GAME_SETTINGS

# ── Constants ─────────────────────────────────────────────────────────────────
N_STUDENTS           = 30
N_ROUNDS             = 10
BASE_PRICE           = float(GAME_SETTINGS.price_per_user)
COST_SCALE           = float(GAME_SETTINGS.cost_scale)
ENV_CAP              = float(GAME_SETTINGS.env_cap)
SOC_CAP              = float(GAME_SETTINGS.social_cap)
DELTA                = float(GAME_SETTINGS.price_sensitivity_delta)
AUDIT_PROB           = 0.30
CATCH_PROB           = 0.80
HIGH_SHARE_THRESHOLD = 0.085
SEED                 = 42

RUNS = [
    {"label": "A", "alpha": 3.0, "beta": 3.0},
    {"label": "B", "alpha": 1.5, "beta": 3.2},
]

OUTPUT_PATH = Path(__file__).parent / "classroom_simulation_compare.xlsx"

# ── Load data ─────────────────────────────────────────────────────────────────
suppliers_df, users_df = get_tables()
N_USERS      = len(users_df)
supplier_ids = suppliers_df["supplier_id"].tolist()

HAS_CATEGORIES = (
    "category" in suppliers_df.columns
    and suppliers_df["category"].notna().any()
)
cat_groups: dict[str, list[str]] = {}
if HAS_CATEGORIES:
    for _, row in suppliers_df.iterrows():
        cat = str(row["category"]).strip()
        if cat and cat.lower() != "nan":
            cat_groups.setdefault(cat, []).append(str(row["supplier_id"]))

print(f"Suppliers: {len(supplier_ids)}  |  Users: {N_USERS}  |  Categorical: {HAS_CATEGORIES}")

# ── Supplier helpers ───────────────────────────────────────────────────────────

def avg_attrs(picks: list[str]) -> dict[str, float]:
    sel = suppliers_df[suppliers_df["supplier_id"].isin(picks)]
    if sel.empty:
        return {}
    cols = ["env_risk", "social_risk", "cost_score", "strategic"]
    for c in ("child_labor", "banned_chem"):
        if c in suppliers_df.columns:
            cols.append(c)
    return {c: float(sel[c].mean()) for c in cols if c in sel.columns}


def is_feasible(picks: list[str]) -> bool:
    if not picks:
        return False
    a = avg_attrs(picks)
    return (
        a.get("env_risk", 99)    <= ENV_CAP + 1e-9
        and a.get("social_risk", 99) <= SOC_CAP + 1e-9
        and _category_ok(picks)
    )


def _category_ok(picks: list[str]) -> bool:
    if not HAS_CATEGORIES:
        return True
    pick_set = set(picks)
    return all(
        sum(1 for sid in ids if sid in pick_set) == 1
        for ids in cat_groups.values()
    )


def min_profitable_price(picks: list[str]) -> float:
    a = avg_attrs(picks)
    return COST_SCALE * a.get("cost_score", 0.0) + 0.50


def random_picks(rng: random.Random) -> list[str]:
    if HAS_CATEGORIES:
        return [rng.choice(ids) for ids in cat_groups.values()]
    k = rng.randint(1, min(4, len(supplier_ids)))
    return rng.sample(supplier_ids, k)


def random_feasible_picks(rng: random.Random, max_attempts: int = 300) -> list[str]:
    last = random_picks(rng)
    for _ in range(max_attempts):
        picks = random_picks(rng)
        if is_feasible(picks):
            return picks
        last = picks
    return last


# ── Student agent ─────────────────────────────────────────────────────────────

@dataclass
class Student:
    name:              str
    picks:             list[str]
    price:             float
    prev_profit:       float = 0.0
    prev_demand_share: float = 0.0
    prev_feasible:     bool  = True
    prev_audit_excl:   bool  = False
    played:            bool  = False

    def decide(self, mean_profit: float, rng: random.Random) -> None:
        if not self.played:
            return
        below_mean = self.prev_profit < mean_profit
        high_share = self.prev_demand_share > HIGH_SHARE_THRESHOLD

        if below_mean and not self.prev_feasible:
            self.picks = random_feasible_picks(rng)
            self.price = max(min_profitable_price(self.picks) + 5.0, BASE_PRICE)
            return
        if below_mean and self.prev_audit_excl:
            if rng.random() < 0.50:
                self.picks = random_feasible_picks(rng)
                self.price = max(min_profitable_price(self.picks) + 5.0, self.price)
            return
        if below_mean and self.prev_feasible:
            cut = rng.uniform(0.05, 0.15)
            self.price = max(self.price * (1.0 - cut), min_profitable_price(self.picks))
            return
        if high_share:
            r = rng.random()
            if r < 0.20:
                self.price *= rng.uniform(1.05, 1.15)
            elif r < 0.30:
                self.price = max(
                    self.price * rng.uniform(0.90, 0.95),
                    min_profitable_price(self.picks),
                )


def _make_segments(alpha: float, beta: float) -> list[CustomerSegment]:
    bd = BetaDensity(alpha=alpha, beta=beta)
    us = users_df.sort_values("w_cost").reset_index(drop=True)
    return [
        CustomerSegment(
            segment_id=str(row["user_id"]),
            density=float(bd.density_at((i + 0.5) / N_USERS)),
            w_env=float(row.get("w_env", 0.0)),
            w_social=float(row.get("w_social", 0.0)),
            w_cost=float(row.get("w_cost", 1.0)),
            w_low_quality=float(row.get("w_low_quality", 0.0)),
        )
        for i, (_, row) in enumerate(us.iterrows())
    ]


# ── Run simulation ─────────────────────────────────────────────────────────────

def run_simulation(alpha: float, beta: float, label: str) -> list[pd.DataFrame]:
    """Run a full N_ROUNDS simulation with the given BetaDensity params."""
    rng = random.Random(SEED)

    # Initialise students (same seed → same initial state for every run)
    students: list[Student] = []
    for i in range(1, N_STUDENTS + 1):
        picks = random_picks(rng)
        price = BASE_PRICE + rng.uniform(-10.0, 10.0)
        students.append(Student(name=f"Team_{i:02d}", picks=picks, price=round(price, 2)))

    all_rounds: list[pd.DataFrame] = []

    for round_no in range(1, N_ROUNDS + 1):
        if round_no > 1:
            mean_profit = all_rounds[-1]["realized_profit"].mean()
            for s in students:
                s.decide(mean_profit, rng)

        # Feasible team profiles
        team_profiles: dict[str, dict[str, Any]] = {}
        for s in students:
            if is_feasible(s.picks):
                a = avg_attrs(s.picks)
                team_profiles[s.name] = {
                    "team":             s.name,
                    "price_per_user":   s.price,
                    "picked_suppliers": s.picks[:],
                    "avg_env":          a.get("env_risk",    0.0),
                    "avg_social":       a.get("social_risk", 0.0),
                    "avg_cost":         a.get("cost_score",  0.0),
                    "avg_strategic":    a.get("strategic",   0.0),
                }

        # Audit
        audit_result   = run_audit(
            team_profiles=team_profiles,
            suppliers_df=suppliers_df,
            audit_probability=AUDIT_PROB,
            catch_probability=CATCH_PROB,
            rng=rng,
        )
        audit_excl_set = set(audit_result.excluded_teams)
        for t in audit_excl_set:
            team_profiles.pop(t, None)

        # MNL market
        active_profiles = [
            BuyerProfile(
                team_name=tp["team"],
                price_per_user=tp["price_per_user"],
                avg_env=tp["avg_env"],
                avg_social=tp["avg_social"],
            )
            for tp in team_profiles.values()
        ]
        segments   = _make_segments(alpha, beta)
        mnl_result = run_mnl_market(active_profiles, segments, delta=DELTA, u_outside=None) \
            if active_profiles else None

        rows = []
        for s in students:
            a             = avg_attrs(s.picks)
            feasible      = is_feasible(s.picks)
            audit_excl    = s.name in audit_excl_set
            avg_env       = a.get("env_risk",    0.0)
            avg_social    = a.get("social_risk", 0.0)
            avg_strategic = a.get("strategic",   0.0)

            if not feasible or audit_excl or mnl_result is None:
                demand_share    = 0.0
                realized_profit = 0.0
                market_utility  = 0.0
            else:
                br              = mnl_result.buyer_results.get(s.name)
                demand_share    = br.total_demand if br else 0.0
                cost_per_unit   = COST_SCALE * a.get("cost_score", 0.0)
                realized_profit = demand_share * N_USERS * (s.price - cost_per_unit)
                market_utility  = round((br.realized_utility * N_USERS) if br else 0.0, 3)

            buyer_utility = round(
                0.1 * realized_profit
                + (5.0 - avg_env)
                + (5.0 - avg_social)
                + (5.0 - avg_strategic),
                4,
            )

            rows.append({
                "round_no":         round_no,
                "team":             s.name,
                "picks":            ",".join(s.picks),
                "price":            round(s.price, 2),
                "feasible":         feasible,
                "audit_excluded":   audit_excl,
                "avg_env":          round(avg_env,       3),
                "avg_social":       round(avg_social,    3),
                "avg_cost":         round(a.get("cost_score", 0.0), 3),
                "avg_strategic":    round(avg_strategic, 3),
                "demand_share_pct": round(demand_share * 100, 2),
                "realized_profit":  round(realized_profit, 2),
                "market_utility":   market_utility,
                "buyer_utility":    buyer_utility,
            })

            s.prev_profit       = realized_profit
            s.prev_demand_share = demand_share
            s.prev_feasible     = feasible
            s.prev_audit_excl   = audit_excl
            s.played            = True

        all_rounds.append(pd.DataFrame(rows))

    print(f"Run {label} (alpha={alpha}, beta={beta}) done.")
    return all_rounds


print("\nRunning simulations ...")
results: dict[str, list[pd.DataFrame]] = {}
for cfg in RUNS:
    results[cfg["label"]] = run_simulation(cfg["alpha"], cfg["beta"], cfg["label"])

full: dict[str, pd.DataFrame] = {
    lbl: pd.concat(rds, ignore_index=True) for lbl, rds in results.items()
}

# ── Colour palette ────────────────────────────────────────────────────────────
PALETTE = [
    "#2563eb", "#dc2626", "#059669", "#7c3aed", "#d97706", "#0891b2",
    "#be123c", "#65a30d", "#9333ea", "#ea580c", "#0284c7", "#16a34a",
    "#b91c1c", "#7e22ce", "#ca8a04", "#0369a1", "#15803d", "#c2410c",
    "#6d28d9", "#b45309", "#075985", "#166534", "#991b1b", "#5b21b6",
    "#92400e", "#0c4a6e", "#14532d", "#7f1d1d", "#4c1d95", "#78350f",
]
team_names = [f"Team_{i:02d}" for i in range(1, N_STUDENTS + 1)]
colors     = {t: PALETTE[i % len(PALETTE)] for i, t in enumerate(team_names)}

# ── Excel helpers ──────────────────────────────────────────────────────────────
HDR_A     = PatternFill("solid", fgColor="1e3a5f")   # dark blue  — Run A
HDR_B     = PatternFill("solid", fgColor="064e3b")   # dark green — Run B
HDR_DIFF  = PatternFill("solid", fgColor="4c1d95")   # dark purple — Diff
HDR_FONT  = Font(bold=True, color="FFFFFF")
RED_FILL  = PatternFill("solid", fgColor="FEE2E2")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
POS_FILL  = PatternFill("solid", fgColor="D1FAE5")   # green — positive diff
NEG_FILL  = PatternFill("solid", fgColor="FEE2E2")   # red   — negative diff
CTR       = Alignment(horizontal="center")

ROUND_COLS = [
    "Team", "Picks", "Price", "Feasible", "Audit Excl.",
    "Avg Env", "Avg Social", "Avg Cost", "Strategic",
    "Share %", "Profit", "Mkt Utility", "Buyer Utility",
]
DIFF_COLS = ["Team", "Share % (B-A)", "Profit (B-A)", "Mkt Utility (B-A)", "Buyer Utility (B-A)"]
N_COLS    = len(ROUND_COLS)

def _header(ws, start_row: int, cols: list[str], fill: PatternFill) -> None:
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font, c.fill, c.alignment = HDR_FONT, fill, CTR

def _section_title(ws, row: int, text: str, fill: PatternFill, ncols: int) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, color="FFFFFF", size=11)
    for ci in range(1, ncols + 1):
        ws.cell(row=row, column=ci).fill = fill

def _fig_to_image(fig: plt.Figure, dpi: int = 110) -> XLImage:
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return XLImage(buf)

def _scatter_ax(ax, xvals, yvals, labels, clrs, xlabel, ylabel, title):
    for x, y, lbl, c in zip(xvals, yvals, labels, clrs):
        ax.scatter(x, y, color=c, s=50, zorder=3, edgecolors="white", linewidths=0.4)
        ax.annotate(lbl.replace("Team_", "T"), (x, y), fontsize=5.5,
                    ha="center", va="bottom", xytext=(0, 3), textcoords="offset points")
    ax.set_xlabel(xlabel, fontsize=9); ax.set_ylabel(ylabel, fontsize=9)
    ax.set_title(title, fontsize=9, fontweight="bold"); ax.grid(alpha=0.2)

def _write_run_table(ws, start_row: int, rdf: pd.DataFrame, fill: PatternFill) -> None:
    _header(ws, start_row, ROUND_COLS, fill)
    for ri, (_, row) in enumerate(rdf.iterrows(), start_row + 1):
        ws.cell(ri,  1, row["team"])
        ws.cell(ri,  2, row["picks"])
        ws.cell(ri,  3, row["price"])
        ws.cell(ri,  4, "Yes" if row["feasible"] else "No")
        ws.cell(ri,  5, "Yes" if row["audit_excluded"] else "")
        ws.cell(ri,  6, row["avg_env"])
        ws.cell(ri,  7, row["avg_social"])
        ws.cell(ri,  8, row["avg_cost"])
        ws.cell(ri,  9, row["avg_strategic"])
        ws.cell(ri, 10, row["demand_share_pct"])
        ws.cell(ri, 11, row["realized_profit"])
        ws.cell(ri, 12, row["market_utility"])
        ws.cell(ri, 13, row["buyer_utility"])
        row_fill = (RED_FILL  if not row["feasible"] else
                    WARN_FILL if row["audit_excluded"] else None)
        if row_fill:
            for ci in range(1, N_COLS + 1):
                ws.cell(ri, ci).fill = row_fill

def _write_diff_table(ws, start_row: int, rdf_a: pd.DataFrame, rdf_b: pd.DataFrame) -> None:
    _header(ws, start_row, DIFF_COLS, HDR_DIFF)
    merged = rdf_a[["team", "demand_share_pct", "realized_profit",
                     "market_utility", "buyer_utility"]].copy()
    merged = merged.merge(
        rdf_b[["team", "demand_share_pct", "realized_profit",
               "market_utility", "buyer_utility"]],
        on="team", suffixes=("_a", "_b"),
    )
    for ri, (_, row) in enumerate(merged.iterrows(), start_row + 1):
        d_share  = round(row["demand_share_pct_b"]  - row["demand_share_pct_a"],  2)
        d_profit = round(row["realized_profit_b"]   - row["realized_profit_a"],   2)
        d_mkt    = round(row["market_utility_b"]    - row["market_utility_a"],    3)
        d_buyer  = round(row["buyer_utility_b"]     - row["buyer_utility_a"],     4)

        ws.cell(ri, 1, row["team"])
        ws.cell(ri, 2, d_share)
        ws.cell(ri, 3, d_profit)
        ws.cell(ri, 4, d_mkt)
        ws.cell(ri, 5, d_buyer)

        for ci, val in zip((2, 3, 4, 5), (d_share, d_profit, d_mkt, d_buyer)):
            if val > 0:
                ws.cell(ri, ci).fill = POS_FILL
            elif val < 0:
                ws.cell(ri, ci).fill = NEG_FILL


# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Row offsets per round sheet:
#   Row 1               : Run A section title
#   Row 2               : Run A header
#   Rows 3 .. N+2       : Run A data   (N_STUDENTS rows)
#   Row N+3             : (blank)
#   Row N+4             : Run B section title
#   Row N+5             : Run B header
#   Rows N+6 .. 2N+5    : Run B data
#   Row 2N+6            : (blank)
#   Row 2N+7            : Diff section title
#   Row 2N+8            : Diff header
#   Rows 2N+9 .. 3N+8   : Diff data
#   Row 3N+9            : (blank)
#   Charts start at 3N+11

ROW_A_TITLE = 1
ROW_A_HDR   = 2
ROW_A_DATA  = 3                          # .. ROW_A_DATA + N_STUDENTS - 1

ROW_B_TITLE = ROW_A_DATA + N_STUDENTS + 1
ROW_B_HDR   = ROW_B_TITLE + 1
ROW_B_DATA  = ROW_B_HDR  + 1

ROW_D_TITLE = ROW_B_DATA + N_STUDENTS + 1
ROW_D_HDR   = ROW_D_TITLE + 1
ROW_D_DATA  = ROW_D_HDR  + 1

CHART_START = ROW_D_DATA + N_STUDENTS + 2

for round_no in range(1, N_ROUNDS + 1):
    ws  = wb.create_sheet(title=f"Round {round_no}")
    rdf_a = results["A"][round_no - 1]
    rdf_b = results["B"][round_no - 1]

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for col_letter in ("C","D","E","F","G","H","I","J","K","L","M"):
        ws.column_dimensions[col_letter].width = 13

    # Section A
    _section_title(ws, ROW_A_TITLE,
                   f"Run A  —  alpha={RUNS[0]['alpha']}, beta={RUNS[0]['beta']}  (Round {round_no})",
                   HDR_A, N_COLS)
    _write_run_table(ws, ROW_A_HDR, rdf_a, HDR_A)

    # Section B
    _section_title(ws, ROW_B_TITLE,
                   f"Run B  —  alpha={RUNS[1]['alpha']}, beta={RUNS[1]['beta']}  (Round {round_no})",
                   HDR_B, N_COLS)
    _write_run_table(ws, ROW_B_HDR, rdf_b, HDR_B)

    # Difference section
    _section_title(ws, ROW_D_TITLE,
                   f"Difference  (B minus A)  —  green = B better, red = A better  (Round {round_no})",
                   HDR_DIFF, len(DIFF_COLS))
    _write_diff_table(ws, ROW_D_HDR, rdf_a, rdf_b)

    # ── Charts ────────────────────────────────────────────────────────────────
    act_a = rdf_a[rdf_a["demand_share_pct"] > 0].copy()
    act_b = rdf_b[rdf_b["demand_share_pct"] > 0].copy()

    def _pie(rdf_active, run_label, alpha_val, beta_val):
        if rdf_active.empty:
            return None
        large = rdf_active[rdf_active["demand_share_pct"] >= 2.0]
        small = rdf_active[rdf_active["demand_share_pct"] <  2.0]
        labels = large["team"].tolist()
        sizes  = large["demand_share_pct"].tolist()
        clrs   = [colors[t] for t in labels]
        if not small.empty:
            labels.append(f"Others ({len(small)})")
            sizes.append(float(small["demand_share_pct"].sum()))
            clrs.append("#94a3b8")
        fig, ax = plt.subplots(figsize=(6.5, 5))
        wedges, _, auts = ax.pie(sizes, labels=None,
                                  autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
                                  colors=clrs, startangle=140, pctdistance=0.78)
        for at in auts: at.set_fontsize(7)
        ax.legend(wedges, labels, title="Team", loc="center left",
                  bbox_to_anchor=(1.0, 0.5), fontsize=7, title_fontsize=8)
        ax.set_title(
            f"Run {run_label} (α={alpha_val}, β={beta_val})\n"
            f"Round {round_no} — Market Share  (active: {len(rdf_active)})",
            fontsize=9, fontweight="bold")
        plt.tight_layout()
        return _fig_to_image(fig)

    img_pie_a = _pie(act_a, "A", RUNS[0]["alpha"], RUNS[0]["beta"])
    img_pie_b = _pie(act_b, "B", RUNS[1]["alpha"], RUNS[1]["beta"])
    if img_pie_a:
        img_pie_a.anchor = f"A{CHART_START}"
        ws.add_image(img_pie_a)
    if img_pie_b:
        img_pie_b.anchor = f"M{CHART_START}"
        ws.add_image(img_pie_b)

    # Scatter: Buyer Utility vs Market Utility — A and B side by side
    CHART_ROW2 = CHART_START + 33
    if not act_a.empty and not act_b.empty:
        fig_s, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))
        _scatter_ax(ax1, act_a["market_utility"], act_a["buyer_utility"],
                    act_a["team"], [colors[t] for t in act_a["team"]],
                    "Market Utility", "Buyer Utility",
                    f"Run A (α={RUNS[0]['alpha']}, β={RUNS[0]['beta']}) — Buyer vs Market Utility")
        _scatter_ax(ax2, act_b["market_utility"], act_b["buyer_utility"],
                    act_b["team"], [colors[t] for t in act_b["team"]],
                    "Market Utility", "Buyer Utility",
                    f"Run B (α={RUNS[1]['alpha']}, β={RUNS[1]['beta']}) — Buyer vs Market Utility")
        plt.suptitle(f"Round {round_no} — Buyer Utility vs Market Utility", fontsize=11, fontweight="bold")
        plt.tight_layout()
        img_s1 = _fig_to_image(fig_s)
        img_s1.anchor = f"A{CHART_ROW2}"
        ws.add_image(img_s1)

    # Scatter: Profit vs Buyer Utility
    CHART_ROW3 = CHART_ROW2 + 33
    if not act_a.empty and not act_b.empty:
        fig_s2, (ax3, ax4) = plt.subplots(1, 2, figsize=(13, 5))
        _scatter_ax(ax3, act_a["realized_profit"], act_a["buyer_utility"],
                    act_a["team"], [colors[t] for t in act_a["team"]],
                    "Profit ($)", "Buyer Utility",
                    f"Run A — Profit vs Buyer Utility")
        _scatter_ax(ax4, act_b["realized_profit"], act_b["buyer_utility"],
                    act_b["team"], [colors[t] for t in act_b["team"]],
                    "Profit ($)", "Buyer Utility",
                    f"Run B — Profit vs Buyer Utility")
        plt.suptitle(f"Round {round_no} — Profit vs Buyer Utility", fontsize=11, fontweight="bold")
        plt.tight_layout()
        img_s2 = _fig_to_image(fig_s2)
        img_s2.anchor = f"A{CHART_ROW3}"
        ws.add_image(img_s2)

    # Scatter: Profit vs Market Utility
    CHART_ROW4 = CHART_ROW3 + 33
    if not act_a.empty and not act_b.empty:
        fig_s3, (ax5, ax6) = plt.subplots(1, 2, figsize=(13, 5))
        _scatter_ax(ax5, act_a["realized_profit"], act_a["market_utility"],
                    act_a["team"], [colors[t] for t in act_a["team"]],
                    "Profit ($)", "Market Utility",
                    f"Run A — Profit vs Market Utility")
        _scatter_ax(ax6, act_b["realized_profit"], act_b["market_utility"],
                    act_b["team"], [colors[t] for t in act_b["team"]],
                    "Profit ($)", "Market Utility",
                    f"Run B — Profit vs Market Utility")
        plt.suptitle(f"Round {round_no} — Profit vs Market Utility", fontsize=11, fontweight="bold")
        plt.tight_layout()
        img_s3 = _fig_to_image(fig_s3)
        img_s3.anchor = f"A{CHART_ROW4}"
        ws.add_image(img_s3)


# ── Cumulative sheet ───────────────────────────────────────────────────────────
ws_cum = wb.create_sheet(title="Cumulative")

rounds_axis = list(range(1, N_ROUNDS + 1))

def _pivot(df, col):
    return df.pivot_table(index="round_no", columns="team", values=col, fill_value=0.0)

pivots: dict[str, dict[str, pd.DataFrame]] = {}
for lbl in ("A", "B"):
    df = full[lbl]
    pivots[lbl] = {
        "profit":  _pivot(df, "realized_profit").cumsum(),
        "buyer":   _pivot(df, "buyer_utility").cumsum(),
        "market":  _pivot(df, "market_utility").cumsum(),
        "share":   _pivot(df, "demand_share_pct"),
    }

# Final ranking (Run A profit)
final_rank  = pivots["A"]["profit"].iloc[-1].sort_values(ascending=False)
top3        = set(final_rank.head(3).index)
bot3        = set(final_rank.tail(3).index)
highlighted = top3 | bot3

# ── Cumulative profit table (Run A) ──────────────────────────────────────────
cum_by_team = pivots["A"]["profit"].T.copy()
cum_by_team["Total"] = cum_by_team.iloc[:, -1]
cum_by_team = cum_by_team.sort_values("Total", ascending=False)

table_header = ["Team (Run A)"] + [f"After R{r}" for r in range(1, N_ROUNDS + 1)] + ["Total"]
_header(ws_cum, 1, table_header, HDR_A)
for ri, (team, rd) in enumerate(cum_by_team.iterrows(), 2):
    ws_cum.cell(ri, 1, team)
    for ci, r in enumerate(range(1, N_ROUNDS + 1), 2):
        ws_cum.cell(ri, ci, round(rd.get(r, 0.0), 2))
    ws_cum.cell(ri, N_ROUNDS + 2, round(float(rd["Total"]), 2))

ws_cum.column_dimensions["A"].width = 11
for ci in range(2, N_ROUNDS + 3):
    ws_cum.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 13

CUM_CHART_START = N_STUDENTS + 4

def _overlay_line(pivot_a, pivot_b, ylabel, title, figsize=(13, 6)):
    """Both runs on the same axes: A solid, B dashed."""
    fig, ax = plt.subplots(figsize=figsize)
    for team in pivot_a.columns:
        c    = colors.get(team, "#94a3b8")
        hiA  = team in highlighted
        vals_a = pivot_a[team].tolist()
        vals_b = pivot_b[team].tolist() if team in pivot_b.columns else [0] * N_ROUNDS
        if hiA:
            ax.plot(rounds_axis, vals_a, color=c, linewidth=2.2, alpha=1.0,
                    label=f"{team} A", zorder=3, marker="o", markersize=4)
            ax.plot(rounds_axis, vals_b, color=c, linewidth=1.6, alpha=0.7,
                    linestyle="--", label=f"{team} B", zorder=3, marker="s", markersize=3)
        else:
            ax.plot(rounds_axis, vals_a, color="#94a3b8", linewidth=0.8, alpha=0.3, zorder=1)
            ax.plot(rounds_axis, vals_b, color="#64748b", linewidth=0.6, alpha=0.2,
                    linestyle="--", zorder=1)
    from matplotlib.lines import Line2D
    legend_extra = [
        Line2D([0], [0], color="gray", linewidth=1.5, linestyle="-",  label=f"Run A (α={RUNS[0]['alpha']}, β={RUNS[0]['beta']})"),
        Line2D([0], [0], color="gray", linewidth=1.5, linestyle="--", label=f"Run B (α={RUNS[1]['alpha']}, β={RUNS[1]['beta']})"),
    ]
    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles + legend_extra, labels + [h.get_label() for h in legend_extra],
              fontsize=8, loc="upper left", framealpha=0.9, ncol=2)
    ax.set_xlabel("Round", fontsize=11); ax.set_ylabel(ylabel, fontsize=11)
    ax.set_title(title + "\n(solid = Run A, dashed = Run B  |  top 3 & bottom 3 by Run A profit highlighted)",
                 fontsize=11, fontweight="bold")
    ax.set_xticks(rounds_axis); ax.grid(axis="y", alpha=0.25)
    plt.tight_layout()
    return fig

def _overlay_scatter(cum_x_a, cum_y_a, cum_x_b, cum_y_b, xlabel, ylabel, title):
    fig, ax = plt.subplots(figsize=(9, 6))
    for team in cum_x_a.columns:
        c      = colors.get(team, "#94a3b8")
        is_h   = team in highlighted
        xa, ya = float(cum_x_a[team].iloc[-1]), float(cum_y_a[team].iloc[-1])
        xb, yb = float(cum_x_b.get(team, {}).get(cum_x_b.index[-1], 0)), \
                 float(cum_y_b.get(team, {}).get(cum_y_b.index[-1], 0)) \
                 if team in cum_x_b.columns else (0, 0)
        # handle pandas Series lookup properly
        if team in cum_x_b.columns:
            xb = float(cum_x_b[team].iloc[-1])
        if team in cum_y_b.columns:
            yb = float(cum_y_b[team].iloc[-1])
        sz  = 100 if is_h else 45
        ax.scatter(xa, ya, color=c, s=sz, marker="o", zorder=3, edgecolors="white", linewidths=0.5)
        ax.scatter(xb, yb, color=c, s=sz, marker="s", zorder=3, edgecolors="white",
                   linewidths=0.5, alpha=0.6)
        if is_h:
            short = team.replace("Team_", "T")
            ax.annotate(short + "A", (xa, ya), fontsize=6.5, fontweight="bold",
                        ha="center", va="bottom", xytext=(0, 4), textcoords="offset points")
            ax.annotate(short + "B", (xb, yb), fontsize=6.5, fontweight="bold",
                        ha="center", va="bottom", xytext=(0, 4), textcoords="offset points",
                        color="#374151")
    from matplotlib.lines import Line2D
    ax.legend([
        Line2D([0], [0], marker="o", color="w", markerfacecolor="gray", markersize=8, label=f"Run A (α={RUNS[0]['alpha']}, β={RUNS[0]['beta']})"),
        Line2D([0], [0], marker="s", color="w", markerfacecolor="gray", markersize=8, label=f"Run B (α={RUNS[1]['alpha']}, β={RUNS[1]['beta']})"),
    ], [f"Run A (α={RUNS[0]['alpha']}, β={RUNS[0]['beta']})", f"Run B (α={RUNS[1]['alpha']}, β={RUNS[1]['beta']})"],
        fontsize=9)
    ax.set_xlabel(xlabel, fontsize=11); ax.set_ylabel(ylabel, fontsize=11)
    ax.set_title(title, fontsize=11, fontweight="bold"); ax.grid(alpha=0.2)
    plt.tight_layout()
    return fig

# Chart 1: Cumulative profit overlay
fig1 = _overlay_line(pivots["A"]["profit"], pivots["B"]["profit"],
                     "Cumulative Profit ($)", "Cumulative Profit")
ws_cum.add_image(_fig_to_image(fig1, 130), f"A{CUM_CHART_START}")

# Chart 2: Cumulative buyer utility overlay
fig2 = _overlay_line(pivots["A"]["buyer"], pivots["B"]["buyer"],
                     "Cumulative Buyer Utility", "Cumulative Buyer Utility")
ws_cum.add_image(_fig_to_image(fig2, 130), f"A{CUM_CHART_START + 33}")

# Chart 3: Cumulative market utility overlay
fig3 = _overlay_line(pivots["A"]["market"], pivots["B"]["market"],
                     "Cumulative Market Utility", "Cumulative Market Utility")
ws_cum.add_image(_fig_to_image(fig3, 130), f"A{CUM_CHART_START + 66}")

# Chart 4: Scatter — cumulative Buyer Utility vs Market Utility (A circles, B squares)
fig4 = _overlay_scatter(
    pivots["A"]["market"], pivots["A"]["buyer"],
    pivots["B"]["market"], pivots["B"]["buyer"],
    "Cumulative Market Utility", "Cumulative Buyer Utility",
    "Cumulative: Buyer Utility vs Market Utility  (o=Run A, s=Run B)",
)
ws_cum.add_image(_fig_to_image(fig4, 130), f"A{CUM_CHART_START + 99}")

# Chart 5: Scatter — cumulative Profit vs Buyer Utility
fig5 = _overlay_scatter(
    pivots["A"]["profit"], pivots["A"]["buyer"],
    pivots["B"]["profit"], pivots["B"]["buyer"],
    "Cumulative Profit ($)", "Cumulative Buyer Utility",
    "Cumulative: Profit vs Buyer Utility  (o=Run A, s=Run B)",
)
ws_cum.add_image(_fig_to_image(fig5, 130), f"A{CUM_CHART_START + 132}")

# Chart 6: Scatter — cumulative Profit vs Market Utility
fig6 = _overlay_scatter(
    pivots["A"]["profit"], pivots["A"]["market"],
    pivots["B"]["profit"], pivots["B"]["market"],
    "Cumulative Profit ($)", "Cumulative Market Utility",
    "Cumulative: Profit vs Market Utility  (o=Run A, s=Run B)",
)
ws_cum.add_image(_fig_to_image(fig6, 130), f"A{CUM_CHART_START + 165}")

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")
