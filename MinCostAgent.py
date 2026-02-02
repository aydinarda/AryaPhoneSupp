"""MinCostAgent.py

Max Profit (Max Margin) agent used by the Streamlit UI.

Design goals:
- Keep this file UI-agnostic (no Streamlit imports).
- Load data only from the workbook's first two sheets (preferred), or from
  sheets named 'Supplier' and 'User' if present.
- Provide a clean API for UI.py to call:
    - load_supplier_user_tables(...)
    - Policy
    - MaxProfitConfig
    - MaxProfitAgent

The agent solves:

  max  Σ_{i,u in last_users} margin_i * z_{i,u}

  where margin_i = P - CostProd(i) - PolicyTariff(i)

Subject to:
- Match at most one supplier per user
- Can only match to selected suppliers
- Total matches <= capacity (default 6 for the "last 6 users" setting)
- Utility feasibility: if matched, Utility(i,u) >= min_utility

Notes:
- "Last users" are taken by the order they appear in the User sheet
  (or natural row order if no explicit ordering column exists).
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

# -------------------------
# Optional Gurobi import
# -------------------------
try:
    import gurobipy as gp
    from gurobipy import GRB

    GUROBI_AVAILABLE = True
except Exception:
    gp = None  # type: ignore
    GRB = None  # type: ignore
    GUROBI_AVAILABLE = False


# -------------------------
# Policy
# -------------------------
@dataclass
class Policy:
    # Multipliers (affect perceived supplier attributes inside utility)
    env_mult: float = 1.0
    social_mult: float = 1.0
    cost_mult: float = 1.0
    strategic_mult: float = 1.0
    improvement_mult: float = 1.0
    low_quality_mult: float = 1.0

    # Penalties/tariffs (subtracted from utility and margin)
    child_labor_penalty: float = 0.0
    banned_chem_penalty: float = 0.0

    def clamp_nonnegative(self) -> "Policy":
        for k, v in self.__dict__.items():
            vv = 0.0 if v is None else float(v)
            setattr(self, k, max(0.0, vv))
        return self

    def to_dict(self) -> Dict[str, float]:
        return {k: float(v) for k, v in self.__dict__.items()}


# -------------------------
# Excel loading
# -------------------------
DEFAULT_XLSX_PATH = "Arya_Phones_Supplier_Selection.xlsx"


def _find_sheet_pair(sheetnames: List[str]) -> Tuple[str, str]:
    """Prefer ('Supplier','User') if present, else take first two sheets."""
    lower = {s.lower(): s for s in sheetnames}
    if "supplier" in lower and "user" in lower:
        return lower["supplier"], lower["user"]
    if len(sheetnames) < 2:
        raise ValueError("Workbook must contain at least two sheets for Supplier and User tables.")
    return sheetnames[0], sheetnames[1]


def _normalize_supplier_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Map common header variants to canonical names
    col_map = {
        "supplier": "supplier_id",
        "supplier_id": "supplier_id",
        "supplier id": "supplier_id",
        "id": "supplier_id",
        "environmental risk": "env_risk",
        "env risk": "env_risk",
        "env_risk": "env_risk",
        "social risk": "social_risk",
        "social_risk": "social_risk",
        "cost score": "cost_score",
        "cost": "cost_score",
        "cost_score": "cost_score",
        "strategic importance": "strategic",
        "strategic": "strategic",
        "improvement potential": "improvement",
        "improvement": "improvement",
        "child labor": "child_labor",
        "child_labor": "child_labor",
        "banned chemicals": "banned_chem",
        "banned chemical": "banned_chem",
        "banned_chem": "banned_chem",
        "low product quality": "low_quality",
        "low quality": "low_quality",
        "low_quality": "low_quality",
    }

    df2 = df.copy()
    df2.columns = [str(c).strip() for c in df2.columns]
    rename = {}
    for c in df2.columns:
        key = str(c).strip().lower()
        if key in col_map:
            rename[c] = col_map[key]
    df2 = df2.rename(columns=rename)

    required = [
        "supplier_id",
        "env_risk",
        "social_risk",
        "cost_score",
        "strategic",
        "improvement",
        "child_labor",
        "banned_chem",
        "low_quality",
    ]
    missing = [c for c in required if c not in df2.columns]
    if missing:
        raise ValueError(f"Supplier sheet is missing columns: {missing}")

    df2 = df2[required].copy()
    df2["supplier_id"] = df2["supplier_id"].astype(str)

    num_cols = [c for c in required if c != "supplier_id"]
    for c in num_cols:
        df2[c] = pd.to_numeric(df2[c], errors="coerce").fillna(0.0).astype(float)

    return df2


def _normalize_user_columns(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {
        "user": "user_id",
        "users": "user_id",
        "user_id": "user_id",
        "user id": "user_id",
        "environmental risk": "w_env",
        "env risk": "w_env",
        "w_env": "w_env",
        "social risk": "w_social",
        "w_social": "w_social",
        "cost score": "w_cost",
        "cost": "w_cost",
        "w_cost": "w_cost",
        "strategic importance": "w_strategic",
        "strategic": "w_strategic",
        "w_strategic": "w_strategic",
        "improvement potential": "w_improvement",
        "improvement": "w_improvement",
        "w_improvement": "w_improvement",
        "low product quality": "w_low_quality",
        "low quality": "w_low_quality",
        "w_low_quality": "w_low_quality",
    }

    df2 = df.copy()
    df2.columns = [str(c).strip() for c in df2.columns]
    rename = {}
    for c in df2.columns:
        key = str(c).strip().lower()
        if key in col_map:
            rename[c] = col_map[key]
    df2 = df2.rename(columns=rename)

    required = ["user_id", "w_env", "w_social", "w_cost", "w_strategic", "w_improvement", "w_low_quality"]
    missing = [c for c in required if c not in df2.columns]
    if missing:
        raise ValueError(f"User sheet is missing columns: {missing}")

    df2 = df2[required].copy()
    df2["user_id"] = df2["user_id"].astype(str)

    for c in required:
        if c == "user_id":
            continue
        df2[c] = pd.to_numeric(df2[c], errors="coerce").fillna(0.0).astype(float)

    # Low quality is a "bad" attribute: larger value should reduce utility
    # If user gave positive weight, we flip it here so it subtracts in utility.
    df2["w_low_quality"] = -df2["w_low_quality"].astype(float)
    return df2


def load_supplier_user_tables(
    xlsx_path: str = DEFAULT_XLSX_PATH,
    supplier_sheet: Optional[str] = None,
    user_sheet: Optional[str] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Loads supplier and user tables from Excel.

    Preference:
      1) If supplier_sheet and user_sheet passed -> use them
      2) If workbook contains sheets named 'Supplier' and 'User' -> use them
      3) Else -> use the first two sheets in workbook order
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    s_sheet, u_sheet = _find_sheet_pair(wb.sheetnames)

    s_name = supplier_sheet or s_sheet
    u_name = user_sheet or u_sheet

    # Read via openpyxl to avoid engine inconsistencies
    def _sheet_to_df(ws: "openpyxl.worksheet.worksheet.Worksheet") -> pd.DataFrame:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()
        # find header row: first row with >=2 non-null entries
        header_idx = 0
        for i, row in enumerate(rows[:30]):
            nn = sum(v is not None and str(v).strip() != "" for v in row)
            if nn >= 2:
                header_idx = i
                break
        header = [str(x).strip() if x is not None else "" for x in rows[header_idx]]
        data = rows[header_idx + 1 :]
        df = pd.DataFrame(data, columns=header)
        # drop completely empty columns
        df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]
        # drop completely empty rows
        df = df.dropna(how="all")
        return df

    ws_s = wb[s_name]
    ws_u = wb[u_name]

    suppliers_raw = _sheet_to_df(ws_s)
    users_raw = _sheet_to_df(ws_u)

    suppliers_df = _normalize_supplier_columns(suppliers_raw)
    users_df = _normalize_user_columns(users_raw)

    return suppliers_df, users_df


# -------------------------
# Max Profit (Max Margin) Agent
# -------------------------
@dataclass
class MaxProfitConfig:
    # Matching setting: 10 users total, optimize only last 6 by default
    last_n_users: int = 6
    capacity: int = 6  # total matches allowed for this retailer
    suppliers_to_select: int = 1

    # Economics
    price_per_match: float = 100.0  # P

    # Feasibility guardrail
    min_utility: float = 0.0

    # Solver controls
    output_flag: int = 0
    big_m: Optional[float] = None


class MaxProfitAgent:
    """
    Max Profit agent: select K suppliers and match the last N users (<= capacity)
    to maximize total margin under a policy.

    Objective:
        max Σ_{i,u} (P - CostProd(i) - Tariff(i)) * z[i,u]
    """

    def __init__(self, suppliers_df: pd.DataFrame, users_df: pd.DataFrame, policy: Policy, cfg: MaxProfitConfig):
        if not GUROBI_AVAILABLE:
            raise RuntimeError("gurobipy is not available. Add it to requirements.txt and ensure a valid license.")

        self.suppliers = suppliers_df.copy()
        self.users = users_df.copy()
        self.policy = policy.clamp_nonnegative()
        self.cfg = cfg

        self.model: Optional["gp.Model"] = None
        self.y = None
        self.z = None
        self._selected_users: List[str] = []

        self._prep()

    def _prep(self) -> None:
        s_req = {"supplier_id", "env_risk", "social_risk", "cost_score", "strategic", "improvement", "child_labor", "banned_chem", "low_quality"}
        u_req = {"user_id", "w_env", "w_social", "w_cost", "w_strategic", "w_improvement", "w_low_quality"}

        if not s_req.issubset(self.suppliers.columns):
            raise ValueError(f"suppliers_df missing columns: {sorted(s_req - set(self.suppliers.columns))}")
        if not u_req.issubset(self.users.columns):
            raise ValueError(f"users_df missing columns: {sorted(u_req - set(self.users.columns))}")

        self.suppliers["supplier_id"] = self.suppliers["supplier_id"].astype(str)
        self.users["user_id"] = self.users["user_id"].astype(str)

        # last N users by sheet order
        n = max(0, int(self.cfg.last_n_users))
        self._selected_users = self.users["user_id"].tolist()[-n:] if n > 0 else self.users["user_id"].tolist()

    def _auto_big_m(self) -> float:
        p = self.policy
        s = self.suppliers
        u = self.users[self.users["user_id"].isin(self._selected_users)].copy()

        # conservative magnitude estimate
        max_s = {
            "env_risk": float((p.env_mult * s["env_risk"]).abs().max()),
            "social_risk": float((p.social_mult * s["social_risk"]).abs().max()),
            "cost_score": float((p.cost_mult * s["cost_score"]).abs().max()),
            "strategic": float((p.strategic_mult * s["strategic"]).abs().max()),
            "improvement": float((p.improvement_mult * s["improvement"]).abs().max()),
            "low_quality": float((p.low_quality_mult * s["low_quality"]).abs().max()),
        }
        max_u = {
            "w_env": float(u["w_env"].abs().max() if len(u) else 0.0),
            "w_social": float(u["w_social"].abs().max() if len(u) else 0.0),
            "w_cost": float(u["w_cost"].abs().max() if len(u) else 0.0),
            "w_strategic": float(u["w_strategic"].abs().max() if len(u) else 0.0),
            "w_improvement": float(u["w_improvement"].abs().max() if len(u) else 0.0),
            "w_low_quality": float(u["w_low_quality"].abs().max() if len(u) else 0.0),
        }

        pref_max = (
            max_u["w_env"] * max_s["env_risk"]
            + max_u["w_social"] * max_s["social_risk"]
            + max_u["w_cost"] * max_s["cost_score"]
            + max_u["w_strategic"] * max_s["strategic"]
            + max_u["w_improvement"] * max_s["improvement"]
            + max_u["w_low_quality"] * max_s["low_quality"]
        )
        pen_max = float((p.child_labor_penalty * s["child_labor"] + p.banned_chem_penalty * s["banned_chem"]).max())
        return float(pref_max + pen_max + 10.0)

    def build(self, name: str = "MaxProfitAgent") -> "gp.Model":
        cfg = self.cfg
        pol = self.policy

        Suppliers = self.suppliers["supplier_id"].tolist()
        Users = self._selected_users

        # supplier dicts
        s_env = dict(zip(self.suppliers["supplier_id"], self.suppliers["env_risk"]))
        s_social = dict(zip(self.suppliers["supplier_id"], self.suppliers["social_risk"]))
        s_cost = dict(zip(self.suppliers["supplier_id"], self.suppliers["cost_score"]))
        s_str = dict(zip(self.suppliers["supplier_id"], self.suppliers["strategic"]))
        s_imp = dict(zip(self.suppliers["supplier_id"], self.suppliers["improvement"]))
        s_lq = dict(zip(self.suppliers["supplier_id"], self.suppliers["low_quality"]))
        s_child = dict(zip(self.suppliers["supplier_id"], self.suppliers["child_labor"]))
        s_banned = dict(zip(self.suppliers["supplier_id"], self.suppliers["banned_chem"]))

        # user dicts
        udf = self.users[self.users["user_id"].isin(Users)].copy()
        u_env = dict(zip(udf["user_id"], udf["w_env"]))
        u_soc = dict(zip(udf["user_id"], udf["w_social"]))
        u_cost = dict(zip(udf["user_id"], udf["w_cost"]))
        u_str = dict(zip(udf["user_id"], udf["w_strategic"]))
        u_imp = dict(zip(udf["user_id"], udf["w_improvement"]))
        u_lq = dict(zip(udf["user_id"], udf["w_low_quality"]))  # already NEG

        M = cfg.big_m if cfg.big_m is not None else self._auto_big_m()

        m = gp.Model(name)
        m.Params.OutputFlag = int(cfg.output_flag)

        # Vars
        y = m.addVars(Suppliers, vtype=GRB.BINARY, name="y_select")
        z = m.addVars(Suppliers, Users, vtype=GRB.BINARY, name="z_match")

        # Select K suppliers
        m.addConstr(gp.quicksum(y[i] for i in Suppliers) == int(cfg.suppliers_to_select), name="select_k")

        # Each user matched at most once
        for u in Users:
            m.addConstr(gp.quicksum(z[i, u] for i in Suppliers) <= 1, name=f"user_once[{u}]")

        # Linking: can only match to selected suppliers
        for i in Suppliers:
            for u in Users:
                m.addConstr(z[i, u] <= y[i], name=f"link[{i},{u}]")

        # Capacity
        m.addConstr(gp.quicksum(z[i, u] for i in Suppliers for u in Users) <= int(cfg.capacity), name="capacity")

        # Utility feasibility
        for i in Suppliers:
            for u in Users:
                user_pref = (
                    (u_env[u] * (pol.env_mult * s_env[i]))
                    + (u_soc[u] * (pol.social_mult * s_social[i]))
                    + (u_cost[u] * (pol.cost_mult * s_cost[i]))
                    + (u_str[u] * (pol.strategic_mult * s_str[i]))
                    + (u_imp[u] * (pol.improvement_mult * s_imp[i]))
                    + (u_lq[u] * (pol.low_quality_mult * s_lq[i]))
                )
                policy_penalty = (pol.child_labor_penalty * s_child[i]) + (pol.banned_chem_penalty * s_banned[i])
                utility = user_pref - policy_penalty
                m.addConstr(utility >= float(cfg.min_utility) - M * (1 - z[i, u]), name=f"utility[{i},{u}]")

        # Margin per supplier
        cost_prod = {i: float(pol.cost_mult * s_cost[i]) for i in Suppliers}
        tariff_i = {i: float(pol.child_labor_penalty * s_child[i] + pol.banned_chem_penalty * s_banned[i]) for i in Suppliers}
        margin_i = {i: float(cfg.price_per_match - cost_prod[i] - tariff_i[i]) for i in Suppliers}

        # Objective
        Z_margin = gp.quicksum(margin_i[i] * z[i, u] for i in Suppliers for u in Users)
        m.setObjective(Z_margin, GRB.MAXIMIZE)

        self.model, self.y, self.z = m, y, z
        return m

    def solve(self) -> Dict[str, Any]:
        if self.model is None:
            self.build()

        assert self.model is not None
        self.model.optimize()

        if self.model.Status not in (GRB.OPTIMAL, GRB.SUBOPTIMAL):
            raise RuntimeError(f"No solution. Status={self.model.Status}")

        Suppliers = self.suppliers["supplier_id"].tolist()
        Users = self._selected_users

        chosen = [i for i in Suppliers if self.y[i].X > 0.5]
        pairs: List[Tuple[str, str]] = [(u, i) for i in Suppliers for u in Users if self.z[i, u].X > 0.5]

        df = pd.DataFrame(pairs, columns=["user_id", "supplier_id"])
        if len(df):
            # compute margin & utility for reporting
            pol = self.policy
            s = self.suppliers.set_index("supplier_id")
            u = self.users.set_index("user_id")

            df["tariff"] = df["supplier_id"].map(lambda sid: float(pol.child_labor_penalty * s.loc[sid, "child_labor"] + pol.banned_chem_penalty * s.loc[sid, "banned_chem"]))
            df["cost_prod"] = df["supplier_id"].map(lambda sid: float(pol.cost_mult * s.loc[sid, "cost_score"]))
            df["margin"] = float(self.cfg.price_per_match) - df["cost_prod"] - df["tariff"]

            def _utility(row: pd.Series) -> float:
                sid = row["supplier_id"]
                uid = row["user_id"]
                user_pref = (
                    u.loc[uid, "w_env"] * (pol.env_mult * s.loc[sid, "env_risk"])
                    + u.loc[uid, "w_social"] * (pol.social_mult * s.loc[sid, "social_risk"])
                    + u.loc[uid, "w_cost"] * (pol.cost_mult * s.loc[sid, "cost_score"])
                    + u.loc[uid, "w_strategic"] * (pol.strategic_mult * s.loc[sid, "strategic"])
                    + u.loc[uid, "w_improvement"] * (pol.improvement_mult * s.loc[sid, "improvement"])
                    + u.loc[uid, "w_low_quality"] * (pol.low_quality_mult * s.loc[sid, "low_quality"])
                )
                penalty = float(pol.child_labor_penalty * s.loc[sid, "child_labor"] + pol.banned_chem_penalty * s.loc[sid, "banned_chem"])
                return float(user_pref - penalty)

            df["utility"] = df.apply(_utility, axis=1)
            df = df.sort_values(["supplier_id", "user_id"]).reset_index(drop=True)
        else:
            df["tariff"] = []
            df["cost_prod"] = []
            df["margin"] = []
            df["utility"] = []

        return {
            "status": int(self.model.Status),
            "objective_value": float(self.model.ObjVal),
            "chosen_suppliers": chosen,
            "selected_users": Users,
            "num_matched": int(len(pairs)),
            "matches": df,
            "policy": self.policy.to_dict(),
            "cfg": self.cfg,
        }
